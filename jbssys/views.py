import csv
import re
import threading
from io import BytesIO
from uuid import uuid4

import pandas as pd
import xlwt
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.db.models import Q

from django.http import HttpResponse, HttpResponseRedirect
from django.utils.encoding import smart_str
from openpyxl import load_workbook
from requests import Response
from rest_framework.decorators import api_view

from .decorators import user_is_role_adm
from .googledriveprocces import GoogleDrive
from .models import (
    DataFormTable,
    Pacient,
    gen_aaaannnna,
    UserProfile,
    User,
    ProfileFormTable,
    Drug,
    MedicalOperation,
    ChldPacientDurBreast,
    RelationhipPacient,
    Cancer,
    CancerHistory,
    DataDictionary,
)

import jbssys.validators as vaidator_data


def create_f_score_data_dict():
    f_score_data_dict = {}

    f_score_data_dict_db = DataDictionary.objects.all()

    for data_ in f_score_data_dict_db:
        key1_lower = data_.f_code
        key2_lower = data_.value.lower().strip()
        try:


            f_score_data_dict[key1_lower][key2_lower] = data_.f_score
        except:
            f_score_data_dict[key1_lower] = {key2_lower:data_.f_score}


    return f_score_data_dict



def master_tableLoadFile(request):
    if request.method =='GET':

        return

def pacientDownLoadFile(request):
    pass


def master_tableDownloadFileTemps(request):
    if request.method == "GET":
        columns_names1 = ['Create', 'First Name', 'Surname', 'Date of birth', 'Client phone',
                         'NHS ID', 'Screener', 'Datapoint 8', 'Datapoint 9', 'Datapoint 10',
                         'Datapoint 11', 'Datapoint 12', 'Datapoint 13', 'Datapoint 14', 'Datapoint 15',
                         'Datapoint 16', 'Datapoint 16A', 'Datapoint 16B', 'Datapoint 17', 'Datapoint 18',
                         'Datapoint 19', 'Datapoint 20',
                         'Datapoint 21', 'Datapoint 22', 'Datapoint 23', 'Datapoint 24', 'Datapoint 25',
                         'Datapoint 26', 'Datapoint 27', 'Datapoint 28', 'Datapoint 29', 'Datapoint 30',
                         'Datapoint 31', 'Datapoint 32', 'Datapoint 33', 'Datapoint 34', 'Datapoint 35',
                         'Datapoint 36', 'Datapoint 37', 'Datapoint 38', 'Datapoint 39', 'Datapoint 40',
                         'Datapoint 41', 'Datapoint 42', 'Datapoint 43', 'Datapoint 44', 'Datapoint 45',
                         'Datapoint 46', 'Datapoint 47', 'Datapoint 48', 'Datapoint 49', 'Datapoint 50',
                         'Datapoint 51', 'Datapoint 52', 'Datapoint 53', 'Datapoint 54', 'Datapoint 55',
                         'Datapoint 56', 'Datapoint 57', 'Datapoint 58', 'Datapoint 59', 'Datapoint 60',
                         'Datapoint 61', 'Datapoint 62', 'Datapoint 63', 'Datapoint 64', 'Datapoint 65',
                         'Datapoint 66', 'Datapoint 67', 'Datapoint 68', 'Datapoint 69', 'Datapoint 70',
                         'Datapoint 71', 'Datapoint 72', 'Datapoint 73', 'Datapoint 74', 'Datapoint 75',
                         'Datapoint 76', 'Datapoint 77', 'Datapoint 78', 'Datapoint 79', 'Datapoint 80',
                         'Datapoint 81', 'Datapoint 82', 'Datapoint 83', 'Datapoint 84', 'Datapoint 85',
                         'Datapoint 86', 'Datapoint 87', 'Datapoint 88', 'Datapoint 89', 'Datapoint 90',
                         'Datapoint 91', 'Datapoint 92', 'Datapoint 93', 'Datapoint 94', 'Datapoint 95',
                         'Datapoint 96', 'Datapoint 97', 'Datapoint 98', 'Datapoint 99', 'Datapoint 100',
                         'Datapoint 101', 'Datapoint 102', 'Datapoint 103', 'Datapoint 104', 'Datapoint 105',
                         'Datapoint 106', 'Datapoint 107', 'Datapoint 108', 'Datapoint 109', 'Datapoint 110',
                         'Datapoint 111', 'Datapoint 112', 'Datapoint 113', 'Datapoint 114', 'Datapoint 115',
                         'Datapoint 116', 'Datapoint 117', 'Datapoint 118', 'Datapoint 119','GD id']

        columns_names2 = ['Create', 'First Name', 'Surname', 'Date of birth', 'Client phone',
                          'NHS ID', 'Datapoint 8', 'Datapoint 8', 'Datapoint 9', 'Datapoint 10',
                          'Datapoint 11', 'Datapoint 12', 'Datapoint 13a', 'Datapoint 13b', 'Datapoint 14', 'Datapoint 15',
                          'Datapoint 16', 'Datapoint 17', 'Datapoint 18','Datapoint 19', 'Datapoint 20',
                          'Datapoint 21', 'Datapoint 22', 'Datapoint 23', 'Datapoint 24', 'Datapoint 25',
                          'Datapoint 26', 'Datapoint 27', 'Datapoint 28', 'Datapoint 29', 'Datapoint 30',
                          'Datapoint 31', 'Datapoint 32', 'Datapoint 33', 'Datapoint 34', 'Datapoint 35',
                          'Datapoint 36', 'Datapoint 37', 'Datapoint 38', 'Datapoint 39', 'Datapoint 40',
                          'Datapoint 41', 'Datapoint 42', 'Datapoint 43', 'Datapoint 44', 'Datapoint 45',
                          'Datapoint 46', 'Datapoint 47', 'Datapoint 48', 'Datapoint 49', 'Datapoint 50',
                          'Datapoint 51', 'Datapoint 52', 'Datapoint 53', 'Datapoint 54', 'Datapoint 55',
                          'Datapoint 56', 'Datapoint 57', 'Datapoint 58', 'Datapoint 59', 'Datapoint 60',
                          'Datapoint 61', 'Datapoint 62', 'Datapoint 63', 'Datapoint 64', 'Datapoint 65',
                          'Datapoint 66', 'Datapoint 67', 'Datapoint 68', 'Datapoint 69', 'Datapoint 70',
                          'Datapoint 71', 'Datapoint 72', 'Datapoint 73', 'Datapoint 74', 'Datapoint 75',
                          'Datapoint 76', 'Datapoint 77', 'Datapoint 78', 'Datapoint 79', 'Datapoint 80',
                          'Datapoint 81', 'Datapoint 82', 'Datapoint 83', 'Datapoint 84', 'Datapoint 85',
                          'Datapoint 86', 'Datapoint 87', 'Datapoint 88', 'Datapoint 89', 'Datapoint 90',
                          'Datapoint 91', 'Datapoint 92', 'Datapoint 93', 'Datapoint 94', 'Datapoint 95','GD id']


        response = HttpResponse(content_type='application/force-download')

        response['Content-Disposition'] = 'attachment; filename=%s' % 'Data-FormTemp.xlsx'

        response['X-Sendfile'] = 'static/sample_template/Data-FormTemp.xlsx'

        return response

def master_tableDownloadFile(request):
    if request.method =='GET':
        format_file = request.GET.get('format_file')

        format_file = format_file if format_file else 'csv'

        search_param = {}

        first_name = request.GET.get('first_name')

        last_name = request.GET.get('last_name')

        phone = request.GET.get('phone')

        email = request.GET.get('email')

        nsh_id = request.GET.get('nsh_id')

        default_city = request.GET.get('city')

        home_addres = request.GET.get('addres')
        birth_from = request.GET.get('birth_from')

        birth_to = request.GET.get('birth_to')

        if email:
            email = vaidator_data.validate_email(str(email))

            search_param['email'] = email

        if nsh_id:
            nsh_id = str(nsh_id)

            search_param['nsh_id'] = nsh_id

        if first_name:

            search_param['first_name'] = first_name

        if last_name:
            search_param['last_name'] = last_name

        search_param['is_hide'] = False

        if phone:
            phone = vaidator_data.validate_phone(str(phone))

            search_param['phone'] = phone

        if default_city:
            default_city = str(default_city)

            search_param['default_city'] = default_city

        if home_addres:
            home_addres = str(home_addres)

            search_param['home_addres'] = home_addres

        # if birth_from and birth_to:
        #     birth_from = vaidator_data.validate_date(str(birth_from))
        #
        #     search_param['date_of_birth__range'] = [birth_from, birth_to]
        #
        # elif birth_from:
        #     birth_from = vaidator_data.validate_date(str(birth_from))
        #
        #     search_param['date_of_birth__range'] = [birth_from, '2023-12-30']
        #
        # elif birth_to:
        #     birth_to = vaidator_data.validate_date(str(birth_to))
        #
        #     search_param['date_of_birth__range'] = ['1900-01-01',birth_to]
        #
        # else:
        #     search_param['date_of_birth__range'] = ['1900-01-01','2023-12-30']
        master_table = DataFormTable.objects.filter(**search_param).order_by('id')


        columns_names = ['Create', 'First Name', 'Surname', 'Date of birth', 'Client phone',
                         'NHS ID', 'Screener', 'Datapoint 8', 'Datapoint 9', 'Datapoint 10',
                         'Datapoint 11',  'Datapoint 12',  'Datapoint 13',  'Datapoint 14', 'Datapoint 15',
                         'Datapoint 16',  'Datapoint 16A', 'Datapoint 16B', 'Datapoint 17', 'Datapoint 18', 'Datapoint 19', 'Datapoint 20',
                         'Datapoint 21',  'Datapoint 22',  'Datapoint 23',  'Datapoint 24', 'Datapoint 25',
                         'Datapoint 26',  'Datapoint 27',  'Datapoint 28',  'Datapoint 29', 'Datapoint 30',
                         'Datapoint 31',  'Datapoint 32',  'Datapoint 33',  'Datapoint 34', 'Datapoint 35',
                         'Datapoint 36',  'Datapoint 37',  'Datapoint 38',  'Datapoint 39', 'Datapoint 40',
                         'Datapoint 41',  'Datapoint 42',  'Datapoint 43',  'Datapoint 44', 'Datapoint 45',
                         'Datapoint 46',  'Datapoint 47',  'Datapoint 48',  'Datapoint 49', 'Datapoint 50',
                         'Datapoint 51',  'Datapoint 52',  'Datapoint 53',  'Datapoint 54', 'Datapoint 55',
                         'Datapoint 56',  'Datapoint 57',  'Datapoint 58',  'Datapoint 59', 'Datapoint 60',
                         'Datapoint 61',  'Datapoint 62',  'Datapoint 63',  'Datapoint 64', 'Datapoint 65',
                         'Datapoint 66',  'Datapoint 67',  'Datapoint 68',  'Datapoint 69', 'Datapoint 70',
                         'Datapoint 71',  'Datapoint 72',  'Datapoint 73',  'Datapoint 74', 'Datapoint 75',
                         'Datapoint 76',  'Datapoint 77',  'Datapoint 78',  'Datapoint 79', 'Datapoint 80',
                         'Datapoint 81',  'Datapoint 82',  'Datapoint 83',  'Datapoint 84', 'Datapoint 85',
                         'Datapoint 86',  'Datapoint 87',  'Datapoint 88',  'Datapoint 89', 'Datapoint 90',
                         'Datapoint 91',  'Datapoint 92',  'Datapoint 93',  'Datapoint 94', 'Datapoint 95',
                         'Datapoint 96',  'Datapoint 97',  'Datapoint 98',  'Datapoint 99', 'Datapoint 100',
                         'Datapoint 101', 'Datapoint 102', 'Datapoint 103', 'Datapoint 104', 'Datapoint 105',
                         'Datapoint 106', 'Datapoint 107', 'Datapoint 108', 'Datapoint 109', 'Datapoint 110',
                         'Datapoint 111', 'Datapoint 112', 'Datapoint 113', 'Datapoint 114', 'Datapoint 115',
                         'Datapoint 116', 'Datapoint 117', 'Datapoint 118', 'Datapoint 119']


        if format_file == 'csv':



            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="clinical_master_table.csv"'

            writer = csv.writer(response)
            writer.writerow(columns_names)

            for master_data in master_table:
                writer.writerow([
                    str(master_data.create_datetime),
                    master_data.pacient.first_name,
                    master_data.pacient.last_name,
                    str(master_data.pacient.date_of_birth),
                    master_data.pacient.phone,
                    master_data.pacient.nsh_id,
                    master_data.datapoint_7,



                    master_data.get_datapoint8_display(),


                    master_data.datapoint_9,

                    master_data.get_datapoint_10_display(),

                    master_data.get_datapoint_11_display(),

                    master_data.get_datapoint_12_display(),


                    master_data.get_datapoint_13_display(),

                    master_data.get_datapoint_14_display(),

                    master_data.get_datapoint15_display(),

                    master_data.get_datapoint16_display(),

                    master_data.get_datapoint_16_b_display(),

                    master_data.get_datapoint_16_b_display(),


                    master_data.get_datapoint_17_display(),

                    master_data.get_datapoint18_display(),
                    master_data.get_datapoint19_display(),

                    master_data.get_datapoint20_display(),


                    master_data.get_datapoint_21_display(),

                    master_data.get_datapoint_22_display(),

                    master_data.get_datapoint23_display(),

                    master_data.datapoint_24,

                    master_data.datapoint_25,

                    master_data.get_datapoint26_display(),

                    master_data.get_datapoint_27_display(),

                    master_data.datapoint_28,

                    master_data.datapoint_29,

                    master_data.get_datapoint30_display(),

                    master_data.get_datapoint_31_display(),


                    master_data.get_datapoint_32_display(),

                    master_data.get_datapoint_33_display(),


                    master_data.get_datapoint_34_display(),

                    master_data.get_datapoint_35_display(),

                    master_data.get_datapoint_36_display(),

                    master_data.get_datapoint_37_display(),

                    master_data.get_datapoint_38_display(),

                    master_data.get_datapoint39_display(),

                    master_data.get_datapoint_40_display(),

                    master_data.get_datapoint_41_display(),

                    master_data.get_datapoint42_display(),

                    master_data.datapoint_43,

                    master_data.datapoint_44,

                    master_data.datapoint_45,

                    master_data.get_datapoint46_display(),
                    master_data.get_datapoint47_display(),
                    master_data.get_datapoint48_display(),
                    master_data.get_datapoint49_display(),
                    master_data.get_datapoint50_display(),

                    master_data.get_datapoint_51_display(),

                    master_data.get_datapoint_52_display(),

                    master_data.get_datapoint_53_display(),
                    master_data.get_datapoint_54_display(),


                    master_data.get_datapoint55_display(),
                    master_data.get_datapoint56_display(),
                    master_data.get_datapoint57_display(),
                    master_data.get_datapoint58_display(),
                    master_data.get_datapoint59_display(),

                    master_data.get_datapoint60_display(),

                    master_data.get_datapoint61_display(),

                    master_data.datapoint_62,

                    master_data.get_datapoint63_display(),
                    master_data.get_datapoint64_display(),

                    master_data.get_datapoint_65_display(),

                    master_data.get_datapoint66_display(),

                    master_data.get_datapoint_67_display(),

                    master_data.get_datapoint_68_display(),

                    master_data.get_datapoint69_display(),
                    master_data.get_datapoint70_display(),
                    master_data.get_datapoint71_display(),


                    master_data.get_datapoint_72_display(),
                    master_data.get_datapoint_73_display(),

                    master_data.get_datapoint74_display(),
                    master_data.get_datapoint75_display(),


                    master_data.get_datapoint76_display(),
                    master_data.get_datapoint77_display(),
                    master_data.get_datapoint78_display(),


                    master_data.get_datapoint79_display(),
                    master_data.get_datapoint80_display(),

                    master_data.get_datapoint81_display(),
                    master_data.get_datapoint82_display(),
                    master_data.get_datapoint83_display(),


                    master_data.get_datapoint_84_display(),

                    master_data.get_datapoint_85_display(),

                    master_data.get_datapoint_86_display(),

                    master_data.get_datapoint87_display(),

                    master_data.datapoint_88,

                    master_data.get_datapoint89_display(),

                    master_data.get_datapoint90_display(),

                    master_data.get_datapoint91_display(),

                    master_data.get_datapoint_92_display(),

                    master_data.get_datapoint_93_display(),

                    master_data.get_datapoint_94_display(),

                    master_data.get_datapoint95_display(),

                    master_data.datapoint_96,

                    master_data.get_datapoint97_display(),

                    master_data.get_datapoint98_display(),

                    master_data.get_datapoint_99_display(),

                    master_data.get_datapoint_100_display(),

                    master_data.get_datapoint_101_display(),

                    master_data.get_datapoint_102_display(),
                    master_data.datapoint_103,
                    master_data.get_datapoint104_display(),
                    master_data.get_datapoint105_display(),
                    master_data.get_datapoint106_display(),
                    master_data.get_datapoint107_display(),
                    master_data.get_datapoint108_display(),
                    master_data.get_datapoint109_display(),
                    master_data.get_datapoint110_display(),
                    master_data.get_datapoint111_display(),
                    master_data.get_datapoint112_display(),

                    master_data.get_datapoint113_display(),

                    master_data.get_datapoint114_display(),

                    master_data.get_datapoint115_display(),

                    master_data.get_datapoint116_display(),

                    master_data.get_datapoint117_display(),

                    master_data.get_datapoint118_display(),

                    master_data.get_datapoint119_display()

                ])

            return response

        elif format_file == 'xls':

            # content-type of response
            response = HttpResponse(content_type='application/ms-excel')

            # decide file name
            response['Content-Disposition'] = 'attachment; filename="clinical_master_table.xls"'

            # creating workbook
            wb = xlwt.Workbook(encoding='utf-8')

            # adding sheet
            ws = wb.add_sheet("sheet1")

            # Sheet header, first row
            row_num = 0

            font_style = xlwt.XFStyle()
            # headers are bold
            font_style.font.bold = True

            # column header names, you can use your own headers here
            columns = columns_names

            # write column headers in sheet
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num], font_style)

            # Sheet body, remaining rows
            font_style = xlwt.XFStyle()

            # get your data, from database or from a text file...
            for master_data in master_table:
                row_num = row_num + 1
                a = iter(range(130))

                ws.write(row_num, a.__next__(), str(master_data.create_datetime),                 font_style)
                ws.write(row_num, a.__next__(), master_data.pacient.first_name,                   font_style)
                ws.write(row_num, a.__next__(), master_data.pacient.last_name,                    font_style)
                ws.write(row_num, a.__next__(), str(master_data.pacient.date_of_birth),           font_style)
                ws.write(row_num, a.__next__(), master_data.pacient.phone,                        font_style)
                ws.write(row_num, a.__next__(), master_data.pacient.nsh_id,                       font_style)
                ws.write(row_num, a.__next__(), master_data.datapoint_7,                          font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint8_display(),             font_style)
                ws.write(row_num, a.__next__(), master_data.datapoint_9,                          font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_10_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_11_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_12_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_13_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_14_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint15_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint16_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_16_b_display(),         font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_16_b_display(),         font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_17_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint18_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint19_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint20_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_21_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_22_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint23_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.datapoint_24,                         font_style)
                ws.write(row_num, a.__next__(), master_data.datapoint_25,                         font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint26_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_27_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.datapoint_28_display(),               font_style)
                ws.write(row_num, a.__next__(), master_data.datapoint_29_display(),               font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint30_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_31_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_32_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_33_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_34_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_35_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_36_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_37_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_38_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint39_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_40_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_41_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint42_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.datapoint_43,                         font_style)
                ws.write(row_num, a.__next__(), master_data.datapoint_44,                         font_style)
                ws.write(row_num, a.__next__(), master_data.datapoint_45,                         font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint46_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint47_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint48_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint49_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint50_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_51_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_52_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_53_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_54_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint55_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint56_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint57_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint58_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint59_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint60_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint61_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.datapoint_62,                         font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint63_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint64_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_65_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint66_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_67_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_68_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint69_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint70_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint71_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_72_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_73_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint74_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint75_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint76_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint77_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint78_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint79_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint80_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint81_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint82_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint83_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_84_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_85_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_86_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint87_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.datapoint_88,                         font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint89_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint90_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint91_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_92_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_93_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_94_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint95_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.datapoint_96,                         font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint97_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint98_display(),            font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_99_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_100_display(),          font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_101_display(),          font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint_102_display(),          font_style)
                ws.write(row_num, a.__next__(), master_data.datapoint_103,                        font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint104_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint105_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint106_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint107_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint108_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint109_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint110_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint111_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint112_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint113_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint114_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint115_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint116_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint117_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint118_display(),           font_style)
                ws.write(row_num, a.__next__(), master_data.get_datapoint119_display(),           font_style)


            wb.save(response)
            return response

        else:
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="clinical_master_table.csv"'

            writer = csv.writer(response)
            writer.writerow(columns_names)

            for pacient_data in master_table:
                writer.writerow([
                    pacient_data.id,
                    pacient_data.nsh_id,
                    pacient_data.first_name,
                    pacient_data.last_name,
                    str(pacient_data.date_of_birth),
                    pacient_data.phone,
                    pacient_data.email,
                    pacient_data.default_city,
                    pacient_data.home_addres,

                ])

            return response

def log_out(request):
    logout(request)
    return redirect('/login/')



def login_view(request):
    if request.method == "POST":
        try:
            request_data = request.POST
            username = request_data.get('username')

            password = request_data.get('password')
            user = authenticate(username=username, password=password)
            login(request, user)
            return redirect('/master/')
        except BaseException:
            return render(request, 'auth/login.html')

    else:
        return render(request, 'auth/login.html')
def datadictionary_view(request):
    if request.method == "GET":
        return render(request,'ddactionary/ddactionary.html')

def loadbusinessbuysale(data_):
    if data_:
        parse_id = str(uuid4())
        for company_ in data_:

            #   name = "БЕРЕЗКА ПЛЮС, ООО",

            try:
                name = str(company_['name']).strip().replace(" ", "")
            except BaseException:
                name = ""

            #  type_company = "ООО",
            try:
                type_company = str(
                    company_['type_company']).strip().replace(
                    " ", "")
            except BaseException:
                type_company = ""

            #  brand = "БЕРЕЗКА ПЛЮС",
            try:
                brand = str(company_['brand']).strip().replace(" ", "")
            except BaseException:
                brand = ""

            #  mail_company = null,
            try:
                mail_company = str(
                    company_['mail_company']).strip().replace(
                    " ", "")
            except BaseException:
                mail_company = ""

            #  inn_company = null,
            try:
                inn_company = str(
                    company_['inn_company']).strip().replace(
                    " ", "")
            except BaseException:
                inn_company = ""

            #  check_date = 02.07.2014",
            try:
                check_date = str(
                    company_['check_date']).strip().replace(
                    " ", "")
            except BaseException:
                check_date = ""

            #  manager_ceo = "Алистратов Олег Александрович",
            try:
                manager_ceo = str(
                    company_['manager_ceo']).strip().replace(
                    " ", "")
            except BaseException:
                manager_ceo = ""

            #  manager_fin = "Алистратов Олег Александрович",
            try:
                manager_fin = str(
                    company_['manager_fin']).strip().replace(
                    " ", "")
            except BaseException:
                manager_fin = ""

            #  q_employ = 6
            try:
                q_employ = str(company_['q_employ']).strip().replace(" ", "")
            except BaseException:
                q_employ = ""

            #  title = ""
            try:
                title = str(company_['title']).strip().replace(" ", "")
            except BaseException:
                title = ""

            #  site = ""
            try:
                site = str(company_['site']).strip().replace(" ", "")
            except BaseException:
                site = ""

            #  adress_fact = "83059 г. Донецк, ул. 8-го Марта, 34а"
            try:
                adress_fact = str(
                    company_['adress_fact']).strip().replace(
                    " ", "")
            except BaseException:
                adress_fact = ""

            #  adress_yure = "83059 г. Донецк, ул. 8-го Марта, 34а"
            try:
                adress_yure = str(
                    company_['adress_yure']).strip().replace(
                    " ", "")
            except BaseException:
                adress_yure = ""

            #  adress_post = ""
            try:
                adress_post = str(
                    company_['adress_post']).strip().replace(
                    " ", "")
            except BaseException:
                adress_post = ""

            #  phone_company = "380506233143"
            try:
                phone_company = str(
                    company_['phone_company']).strip().replace(
                    " ", "")
            except BaseException:
                phone_company = ""

            #  phone_company = "380506233143"
            try:
                phone_ceo = str(company_['phone_ceo']).strip().replace(" ", "")
            except BaseException:
                phone_ceo = ""

            #  phone_fin = "380506233143"
            try:
                phone_fin = str(company_['phone_fin']).strip().replace(" ", "")
            except BaseException:
                phone_fin = ""

            #  phone_fax = "380506233143"
            try:
                phone_fax = str(company_['phone_fax']).strip().replace(" ", "")
            except BaseException:
                phone_fax = ""

            #  phone_contact = "[380506233143]"
            try:
                phone_contact = company_['phone_contact']
            except BaseException:
                phone_contact = []

            #  prod_serv = "Формирование и обработка листового стекла"
            try:
                prod_serv = str(company_['prod_serv']).strip().replace(" ", "")
            except BaseException:
                prod_serv = ""

            #  webpage = "https://www.ua-region.info/35711197"
            try:
                webpage = str(company_['webpage']).strip().replace(" ", "")
            except BaseException:
                webpage = ""

            #  description = "https://www.ua-region.info/35711197"
            try:
                description = str(
                    company_['description']).strip().replace(
                    " ", "")
            except BaseException:
                description = ""

            #  company_logo = "https://www.ua-region.info/35711197"

            try:
                company_logo = str(
                    company_['company_logo']).strip().replace(
                    " ", "")
            except BaseException:
                company_logo = None

            #  person_slug = "dsfgsdfexawe"

            try:
                person_slug = str(
                    company_['person_slug']).strip().replace(
                    " ", "")
            except BaseException:
                person_slug = None

            #  source = ""

            try:
                source = str(company_['source']).strip().replace(" ", "")
            except BaseException:
                source = ""

            #  source = list
            try:
                activities = company_['activities']
            except BaseException:
                activities = []

            #  industry = list
            try:
                industry = company_['industry']
            except BaseException:
                industry = []

            businessBuySale_ = BusinessBuySale.objects.create(
                person_slug=person_slug,
                brand=brand,
                title=title,
                offname=name,
                type=type_company,
                type_company=type_company,
                descrition=description,
                prod_serv=prod_serv,
                webpage=webpage,
                company_logo=company_logo,
                source=source,
                adress_yure=adress_yure,
                adress_fact=adress_fact,
                adress_post=adress_post,
                mail_company=mail_company,
                mail_ceo="",
                mail_contact="",
                mail_info="",
                mail_is_find=False,
                mail_is_send=False,
                mail_is_recived=False,
                site_is_work=False,
                disscus_mail="",
                disscus_phone="",
                disscus_name="",
                disscus_status="",
                manager_ceo=manager_ceo,
                manager_fin=manager_fin,
                mail_link="",
                unique_id=parse_id,
                phone_company=phone_company,
                phone_contact=phone_contact,
                phone_ceo=phone_ceo,
                phone_fin=phone_fin,
                phone_fax=phone_fax,
                q_employ=q_employ,
                inn_company=inn_company,
                site=site)
            businessBuySale_.set_phone_contact(phone_contact)
            businessBuySale_.set_activities(activities)
            businessBuySale_.set_industries(industry)
            businessBuySale_.save()

def procces_pacient(record_list):
    first_name = record_list[1]
    last_name = record_list[2]
    date_of_birth = record_list[3]
    phone = record_list[4]
    email = record_list[7]
    email = email if vaidator_data.validate_email(email) else ""



    date_of_birth = vaidator_data.validate_date(date_of_birth)
    date_of_birth = date_of_birth if date_of_birth else None
    if first_name and last_name:

        pacient, created = Pacient.objects.get_or_create(
            first_name=first_name,
            last_name=last_name,
            date_of_birth=date_of_birth,

        )
    else: return None

    if created:
        try:
            nsh_id_old = Pacient.objects.all().order_by('-id')[1]
            nsh_id_old = str(nsh_id_old.nsh_id)

            nsh_id_new = gen_aaaannnna(nsh_id_old)
        except:
            nsh_id_new = gen_aaaannnna()

        pacient.nsh_id = nsh_id_new
    pacient_id = pacient.id

    pacient.phone = phone
    pacient.email = email

    pacient.save()

    return pacient_id

def parse_datapoint_52(data):
    return_data = []
    datas = data.split('\n')
    for n,data_ in enumerate(datas):
        r = data_.split('[]:')[-1].split("months:")[-1].strip()
        return_data.append([n+1,r])

    return return_data

def parse_operation(data):
    return_data = []
    datas = data.split('\n')
    for data_ in datas:
        r = data_.split(",")
        r = [r_i.split(":")[-1].strip() for r_i in r]


        return_data.append(r)




    return return_data


def parse_datapoint_62(data,size=0):
    return_data = []
    datas = data.split('\n')

    for data_ in datas:
        r = data_.split(",")
        r2 = [r_i.split(":")[-1].strip() for r_i in r]
        if size != 0 and len(r2) > size:
            try:
                rl = []
                for i in range(1, int(len(r2) / int(size) + 1)):
                    r_l = r2[(i - 1) * size:i * size]
                    return_data.append(r_l)
            except:

                r_l.append(len(size-r_l)*"")

        else:
            return_data.append(r2)

    return return_data

@user_is_role_adm
def data_dictionatyDelete(request,req_id):
    if request.method == "POST":
        ddict = DataDictionary.objects.get(id=req_id)
        ddict.is_hide = True
        try:
            old_f_score = float(str(ddict.f_score).strip().replace(" ",""))
        except:
            old_f_score = 0
        ddict.save()

        try:
            data_forms = DataFormTable.objects.all()

        except:
            data_forms = []


        for data_form in data_forms:
            try:

                old_u_score_b = data_form.u_score_b
                u_score_b = old_u_score_b - old_f_score
                data_form.u_score_b =  u_score_b
                data_form.save()
            except:
                continue
        try:
            profile_datas = ProfileFormTable.objects.all()

        except:
            profile_datas = []


        for profile_data in profile_datas:
            old_u_score_b = profile_data.u_score_p
            u_score_b = old_u_score_b - old_f_score
            profile_data.u_score_p = u_score_b
            profile_data.save()


        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))


@login_required
def data_dictionatyEdit(request,req_id):
    if request.method == "POST":

        model_code = request.POST.get('edit_model_code')
        f_code = request.POST.get('edit_f_code')
        f_score = request.POST.get('edit_f_score')

        f_code = str(model_code) + str(f_code)



        f_score_a = request.POST.get('edit_f_score_a')

        f_score_b = request.POST.get('edit_f_score_b')

        data_points = request.POST.get('edit_data_points')

        value = request.POST.get('edit_data_dictionary')

        display_destination = request.POST.get('edit_display_destination')

        u_score = request.POST.get('edit_u_score')

        link_logic = request.POST.get('edit_link_logic')
        if link_logic == "Yes":
            link_logic = True
        else:
            link_logic = False



        color_spect = request.POST.get('edit_color_spect')
        data_dictionaty_ = DataDictionary.objects.get(id=req_id)




        try:
            old_f_score = data_dictionaty_.f_score
            old_f_score = float(str(old_f_score).strip().replace(' ', ''))
        except:
            old_f_score = 0.0


        data_dictionaty_.color               = color_spect
        data_dictionaty_.link_logic          = link_logic
        data_dictionaty_.u_score             = u_score
        data_dictionaty_.display_distenation = display_destination
        data_dictionaty_.data_point          = data_points
        data_dictionaty_.value               = value
        data_dictionaty_.f_score_b           = f_score_b
        data_dictionaty_.f_score_a           = f_score_a
        data_dictionaty_.f_score             = f_score
        data_dictionaty_.f_code              = f_code


        data_dictionaty_.save()

        try:
            data_forms = DataFormTable.objects.all()

        except:
            data_forms = []

        print(f_code)

        for data_form in data_forms:

            old_u_score_b = data_form.u_score_b
            u_score_b = old_u_score_b - old_f_score + f_score
            data_form.u_score_b =  u_score_b
            data_form.save()

        try:
            profile_datas = ProfileFormTable.objects.all()

        except:
            profile_datas = []

        print(f_code)

        for profile_data in profile_datas:
            old_u_score_b = profile_data.u_score_p
            u_score_b = old_u_score_b - old_f_score + f_score
            profile_data.u_score_p = u_score_b
            profile_data.save()

        return redirect('/datadictionary/')

def run_f_scores_reloading_weig():
    f_score_data_dict = create_f_score_data_dict()

    def get_data_score(value,f_code):


        try:
            if str(value) != "" and str(value) != None:
                print('value,f_code:',value,f_code) 
                try:
                    fcode_val = f_score_data_dict[str(f_code)][str(value).lower()]
                    print('fcode_val', fcode_val)
                    fcode_val = round(float(str(fcode_val).strip().replace(" ",'')),2)
                    print('fcode_val',fcode_val)
                    return fcode_val
                except:
                    return 0.0

            else:
                return 0.0

        except:
            return 0.0

    if f_score_data_dict:


        try:
            profile_datas = ProfileFormTable.objects.all()

        except:
            profile_datas = []

        for profile_data in profile_datas:

            try:
                # calculation f_score  by add score of ansfer
                f_score = 0.0

                if 'P7' in  f_score_data_dict.keys(): datapoint_9   = profile_data.sex          ;    f_score += get_data_score(datapoint_9,   'P9' )
                if 'P10' in  f_score_data_dict.keys():datapoint_10  = profile_data.datapoint_10 ;    f_score += get_data_score(datapoint_10,  'P10')
                if 'P11' in  f_score_data_dict.keys():datapoint_11  = profile_data.metic_sys    ;    f_score += get_data_score(datapoint_11,  'P11')
                if 'P12' in  f_score_data_dict.keys():datapoint_12  = profile_data.datapoint_12 ;    f_score += get_data_score(datapoint_12,  'P12')
                if 'P13a' in f_score_data_dict.keys():datapoint_13a = profile_data.datapoint_13a;    f_score += get_data_score(datapoint_13a, 'P13a')
                if 'P13b' in f_score_data_dict.keys():datapoint_13b = profile_data.datapoint_13b;    f_score += get_data_score(datapoint_13b, 'P13b')
                if 'P14' in  f_score_data_dict.keys():datapoint_14  = profile_data.datapoint_14 ;    f_score += get_data_score(datapoint_14,  'P14')
                if 'P15' in  f_score_data_dict.keys():datapoint_15  = profile_data.datapoint_15 ;    f_score += get_data_score(datapoint_15,  'P15')
                if 'P16' in  f_score_data_dict.keys():datapoint_16  = profile_data.datapoint_16 ;    f_score += get_data_score(datapoint_16,  'P16')
                if 'P17' in  f_score_data_dict.keys():datapoint_17  = profile_data.datapoint_17 ;    f_score += get_data_score(datapoint_17,  'P17')
                if 'P18' in  f_score_data_dict.keys():datapoint_18  = profile_data.datapoint_18 ;    f_score += get_data_score(datapoint_18,  'P18')
                if 'P19' in  f_score_data_dict.keys():datapoint_19  = profile_data.datapoint_19 ;    f_score += get_data_score(datapoint_19,  'P19')

                if 'P20' in  f_score_data_dict.keys():datapoint_20  = profile_data.datapoint_20 ;  f_score += get_data_score(datapoint_20,  'P20')
                if 'P21' in  f_score_data_dict.keys():datapoint_21  = profile_data.datapoint_21 ;  f_score += get_data_score(datapoint_21,  'P21')
                if 'P22' in  f_score_data_dict.keys():datapoint_22  = profile_data.datapoint_22 ;  f_score += get_data_score(datapoint_22,  'P22')
                if 'P23' in  f_score_data_dict.keys():datapoint_23  = profile_data.datapoint_23 ;  f_score += get_data_score(datapoint_23,  'P23')
                if 'P24' in  f_score_data_dict.keys():datapoint_24  = profile_data.datapoint_24 ;  f_score += get_data_score(datapoint_24,  'P24')
                if 'P25' in  f_score_data_dict.keys():datapoint_25  = profile_data.datapoint_25 ;  f_score += get_data_score(datapoint_25,  'P25')
                if 'P26' in  f_score_data_dict.keys():datapoint_26  = profile_data.datapoint_26 ;  f_score += get_data_score(datapoint_26,  'P26')
                if 'P27' in  f_score_data_dict.keys():datapoint_27  = profile_data.datapoint_27 ;  f_score += get_data_score(datapoint_27,  'P27')
                if 'P28' in  f_score_data_dict.keys():datapoint_28  = profile_data.datapoint_28 ;  f_score += get_data_score(datapoint_28,  'P28')
                if 'P29' in  f_score_data_dict.keys():datapoint_29  = profile_data.datapoint_29 ;  f_score += get_data_score(datapoint_29,  'P29')
            
                if 'P30' in  f_score_data_dict.keys():datapoint_30  = profile_data.datapoint_30 ;  f_score += get_data_score(datapoint_30,  'P30')
                if 'P31' in  f_score_data_dict.keys():datapoint_31  = profile_data.datapoint_31 ;  f_score += get_data_score(datapoint_31,  'P31')
                if 'P32' in  f_score_data_dict.keys():datapoint_32  = profile_data.datapoint_32 ;  f_score += get_data_score(datapoint_32,  'P32')
                if 'P33' in  f_score_data_dict.keys():datapoint_33  = profile_data.datapoint_33 ;  f_score += get_data_score(datapoint_33,  'P33')
                if 'P34' in  f_score_data_dict.keys():datapoint_34  = profile_data.datapoint_34 ;  f_score += get_data_score(datapoint_34,  'P34')
                if 'P35' in  f_score_data_dict.keys():datapoint_35  = profile_data.datapoint_35 ;  f_score += get_data_score(datapoint_35,  'P35')
                if 'P36' in  f_score_data_dict.keys():datapoint_36  = profile_data.datapoint_36 ;  f_score += get_data_score(datapoint_36,  'P36')
                if 'P37' in  f_score_data_dict.keys():datapoint_37  = profile_data.datapoint_37 ;  f_score += get_data_score(datapoint_37,  'P37')
                if 'P38' in  f_score_data_dict.keys():datapoint_38  = profile_data.datapoint_38 ;  f_score += get_data_score(datapoint_38,  'P38')
                if 'P39' in  f_score_data_dict.keys():datapoint_39  = profile_data.datapoint_39 ;  f_score += get_data_score(datapoint_39,  'P39')
            
                if 'P40' in  f_score_data_dict.keys():datapoint_40  = profile_data.datapoint_40 ;  f_score += get_data_score(datapoint_40,  'P40')
                if 'P41' in  f_score_data_dict.keys():datapoint_41  = profile_data.datapoint_41 ;  f_score += get_data_score(datapoint_41,  'P41')
                if 'P42' in  f_score_data_dict.keys():datapoint_42  = profile_data.datapoint_42 ;  f_score += get_data_score(datapoint_42,  'P42')
                if 'P43' in  f_score_data_dict.keys():datapoint_43  = profile_data.datapoint_43 ;  f_score += get_data_score(datapoint_43,  'P43')
                if 'P44' in  f_score_data_dict.keys():datapoint_44  = profile_data.datapoint_44 ;  f_score += get_data_score(datapoint_44,  'P44')
                if 'P45' in  f_score_data_dict.keys():datapoint_45  = profile_data.datapoint_45 ;  f_score += get_data_score(datapoint_45,  'P45')
                if 'P46' in  f_score_data_dict.keys():datapoint_46  = profile_data.datapoint_46 ;  f_score += get_data_score(datapoint_46,  'P46')
                if 'P47' in  f_score_data_dict.keys():datapoint_47  = profile_data.datapoint_47 ;  f_score += get_data_score(datapoint_47,  'P47')
                if 'P48' in  f_score_data_dict.keys():datapoint_48  = profile_data.datapoint_48 ;  f_score += get_data_score(datapoint_48,  'P48')
                if 'P49' in  f_score_data_dict.keys():datapoint_49  = profile_data.datapoint_49 ;  f_score += get_data_score(datapoint_49,  'P49')
              
                if 'P50' in  f_score_data_dict.keys():datapoint_50  = profile_data.datapoint_50 ;  f_score += get_data_score(datapoint_50,  'P50')
                if 'P51' in  f_score_data_dict.keys():datapoint_51  = profile_data.datapoint_51 ;  f_score += get_data_score(datapoint_51,  'P51')
                if 'P52' in  f_score_data_dict.keys():datapoint_52  = profile_data.datapoint_52 ;  f_score += get_data_score(datapoint_52,  'P52')
                if 'P53' in  f_score_data_dict.keys():datapoint_53  = profile_data.datapoint_53 ;  f_score += get_data_score(datapoint_53,  'P53')
                if 'P54' in  f_score_data_dict.keys():datapoint_54  = profile_data.datapoint_54 ;  f_score += get_data_score(datapoint_54,  'P54')
                if 'P55' in  f_score_data_dict.keys():datapoint_55  = profile_data.datapoint_55 ;  f_score += get_data_score(datapoint_55,  'P55')
                if 'P56' in  f_score_data_dict.keys():datapoint_56  = profile_data.datapoint_56 ;  f_score += get_data_score(datapoint_56,  'P56')
                if 'P57' in  f_score_data_dict.keys():datapoint_57  = profile_data.datapoint_57 ;  f_score += get_data_score(datapoint_57,  'P57')
                if 'P58' in  f_score_data_dict.keys():datapoint_58  = profile_data.datapoint_58 ;  f_score += get_data_score(datapoint_58,  'P58')
                if 'P59' in  f_score_data_dict.keys():datapoint_59  = profile_data.datapoint_59 ;  f_score += get_data_score(datapoint_59,  'P59')
              
                if 'P60' in  f_score_data_dict.keys():datapoint_60  = profile_data.datapoint_60 ;  f_score += get_data_score(datapoint_60,  'P60')
                if 'P61' in  f_score_data_dict.keys():datapoint_61  = profile_data.datapoint_61 ;  f_score += get_data_score(datapoint_61,  'P61')
                if 'P62' in  f_score_data_dict.keys():datapoint_62  = profile_data.datapoint_62 ;  f_score += get_data_score(datapoint_62,  'P62')
                if 'P63' in  f_score_data_dict.keys():datapoint_63  = profile_data.datapoint_63 ;  f_score += get_data_score(datapoint_63,  'P63')
                if 'P64' in  f_score_data_dict.keys():datapoint_64  = profile_data.datapoint_64 ;  f_score += get_data_score(datapoint_64,  'P64')
                if 'P65' in  f_score_data_dict.keys():datapoint_65  = profile_data.datapoint_65 ;  f_score += get_data_score(datapoint_65,  'P65')
                if 'P66' in  f_score_data_dict.keys():datapoint_66  = profile_data.datapoint_66 ;  f_score += get_data_score(datapoint_66,  'P66')
                if 'P67' in  f_score_data_dict.keys():datapoint_67  = profile_data.datapoint_67 ;  f_score += get_data_score(datapoint_67,  'P67')
                if 'P68' in  f_score_data_dict.keys():datapoint_68  = profile_data.datapoint_68 ;  f_score += get_data_score(datapoint_68,  'P68')
                if 'P69' in  f_score_data_dict.keys():datapoint_69  = profile_data.datapoint_69 ;  f_score += get_data_score(datapoint_69,  'P69')
              
                if 'P70' in  f_score_data_dict.keys():datapoint_70  = profile_data.datapoint_70 ;  f_score += get_data_score(datapoint_70,  'P70')
                if 'P71' in  f_score_data_dict.keys():datapoint_71  = profile_data.datapoint_71 ;  f_score += get_data_score(datapoint_71,  'P71')
                if 'P72' in  f_score_data_dict.keys():datapoint_72  = profile_data.datapoint_72 ;  f_score += get_data_score(datapoint_72,  'P72')
                if 'P73' in  f_score_data_dict.keys():datapoint_73  = profile_data.datapoint_73 ;  f_score += get_data_score(datapoint_73,  'P73')
                if 'P74' in  f_score_data_dict.keys():datapoint_74  = profile_data.datapoint_74 ;  f_score += get_data_score(datapoint_74,  'P74')
                if 'P75' in  f_score_data_dict.keys():datapoint_75  = profile_data.datapoint_75 ;  f_score += get_data_score(datapoint_75,  'P75')
                if 'P76' in  f_score_data_dict.keys():datapoint_76  = profile_data.datapoint_76 ;  f_score += get_data_score(datapoint_76,  'P76')
                if 'P77' in  f_score_data_dict.keys():datapoint_77  = profile_data.datapoint_77 ;  f_score += get_data_score(datapoint_77,  'P77')
                if 'P78' in  f_score_data_dict.keys():datapoint_78  = profile_data.datapoint_78 ;  f_score += get_data_score(datapoint_78,  'P78')
                if 'P79' in  f_score_data_dict.keys():datapoint_79  = profile_data.datapoint_79 ;  f_score += get_data_score(datapoint_79,  'P79')
              
                if 'P80' in  f_score_data_dict.keys():datapoint_80  = profile_data.datapoint_80 ;  f_score += get_data_score(datapoint_80,  'P80')
                if 'P81' in  f_score_data_dict.keys():datapoint_81  = profile_data.datapoint_81 ;  f_score += get_data_score(datapoint_81,  'P81')
                if 'P82' in  f_score_data_dict.keys():datapoint_82  = profile_data.datapoint_82 ;  f_score += get_data_score(datapoint_82,  'P82')
                if 'P83' in  f_score_data_dict.keys():datapoint_83  = profile_data.datapoint_83 ;  f_score += get_data_score(datapoint_83,  'P83')
                if 'P84' in  f_score_data_dict.keys():datapoint_84  = profile_data.datapoint_84 ;  f_score += get_data_score(datapoint_84,  'P84')
                if 'P85' in  f_score_data_dict.keys():datapoint_85  = profile_data.datapoint_85 ;  f_score += get_data_score(datapoint_85,  'P85')
                if 'P86' in  f_score_data_dict.keys():datapoint_86  = profile_data.datapoint_86 ;  f_score += get_data_score(datapoint_86,  'P86')
                if 'P87' in  f_score_data_dict.keys():datapoint_87  = profile_data.datapoint_87 ;  f_score += get_data_score(datapoint_87,  'P87')
                if 'P88' in  f_score_data_dict.keys():datapoint_88  = profile_data.datapoint_88 ;  f_score += get_data_score(datapoint_88,  'P88')
                if 'P89' in  f_score_data_dict.keys():datapoint_89  = profile_data.datapoint_89 ;  f_score += get_data_score(datapoint_89,  'P89')
              
                if 'P90' in  f_score_data_dict.keys():datapoint_90  = profile_data.datapoint_90 ;  f_score += get_data_score(datapoint_90,  'P90')
                if 'P91' in  f_score_data_dict.keys():datapoint_91  = profile_data.datapoint_91 ;  f_score += get_data_score(datapoint_91,  'P91')
                if 'P92' in  f_score_data_dict.keys():datapoint_92  = profile_data.datapoint_92 ;  f_score += get_data_score(datapoint_92,  'P92')
                if 'P93' in  f_score_data_dict.keys():datapoint_93  = profile_data.datapoint_93 ;  f_score += get_data_score(datapoint_93,  'P93')
                if 'P94' in  f_score_data_dict.keys():datapoint_94  = profile_data.datapoint_94 ;  f_score += get_data_score(datapoint_94,  'P94')
                if 'P95' in  f_score_data_dict.keys():datapoint_95  = profile_data.datapoint_95 ;  f_score += get_data_score(datapoint_95,  'P95')



                profile_data.u_score_p = f_score
                profile_data.save()


            except: continue




        try:
            data_forms = DataFormTable.objects.all()

        except:
            data_forms = []

        for data_form in data_forms:
            try:



                # calculation f_score  by add score of ansfer
                f_score = 0.0

                if 'F7' in  f_score_data_dict.keys():datapoint_7    = data_form.datapoint_7   ;   f_score += get_data_score(datapoint_7 ,   'F7'  )
                if 'F8' in  f_score_data_dict.keys():datapoint_8    = data_form.datapoint_8   ;   f_score += get_data_score(datapoint_8 ,   'F8'  )
                if 'F9' in  f_score_data_dict.keys():datapoint_9    = data_form.datapoint_9   ;   f_score += get_data_score(datapoint_9 ,   'F9'  )
                if 'F10'in  f_score_data_dict.keys():datapoint_10   = data_form.datapoint_10  ;   f_score += get_data_score(datapoint_10,   'F10' )

                if 'F11' in  f_score_data_dict.keys():datapoint_11   = data_form.datapoint_11  ;   f_score += get_data_score(datapoint_11 ,  'F11')
                if 'F12' in  f_score_data_dict.keys():datapoint_12   = data_form.datapoint_12  ;   f_score += get_data_score(datapoint_12 ,  'F12')
                if 'F13' in  f_score_data_dict.keys():datapoint_13   = data_form.datapoint_13  ;   f_score += get_data_score(datapoint_13 ,  'F13')
                if 'F14' in  f_score_data_dict.keys():datapoint_14   = data_form.datapoint_14  ;   f_score += get_data_score(datapoint_14 ,  'F14')
                if 'F15' in  f_score_data_dict.keys():datapoint_15   = data_form.datapoint_15  ;   f_score += get_data_score(datapoint_15 ,  'F15')
                if 'F16' in  f_score_data_dict.keys():datapoint_16   = data_form.datapoint_16  ;   f_score += get_data_score(datapoint_16 ,  'F16')
                if 'F16a'in  f_score_data_dict.keys():datapoint_16_a = data_form.datapoint_16_a;   f_score += get_data_score(datapoint_16_a, 'F16a')
                if 'F16b'in  f_score_data_dict.keys():datapoint_16_b = data_form.datapoint_16_b;   f_score += get_data_score(datapoint_16_b, 'F16b')
                if 'F17' in  f_score_data_dict.keys():datapoint_17   = data_form.datapoint_17  ;   f_score += get_data_score(datapoint_17 ,  'F17')
                if 'F18' in  f_score_data_dict.keys():datapoint_18   = data_form.datapoint_18  ;   f_score += get_data_score(datapoint_18 ,  'F18')
                if 'F19' in  f_score_data_dict.keys():datapoint_19   = data_form.datapoint_19  ;   f_score += get_data_score(datapoint_19 ,  'F19')

                if 'F20' in  f_score_data_dict.keys():datapoint_20   = data_form.datapoint_20  ;   f_score += get_data_score(datapoint_20,   'F20')
                if 'F21' in  f_score_data_dict.keys():datapoint_21   = data_form.datapoint_21  ;   f_score += get_data_score(datapoint_21,   'F21')
                if 'F22' in  f_score_data_dict.keys():datapoint_22   = data_form.datapoint_22  ;   f_score += get_data_score(datapoint_22,   'F22')
                if 'F23' in  f_score_data_dict.keys():datapoint_23   = data_form.datapoint_23  ;   f_score += get_data_score(datapoint_23,   'F23')
                if 'F24' in  f_score_data_dict.keys():datapoint_24   = data_form.datapoint_24  ;   f_score += get_data_score(datapoint_24,   'F24')
                if 'F25' in  f_score_data_dict.keys():datapoint_25   = data_form.datapoint_25  ;   f_score += get_data_score(datapoint_25,   'F25')
                if 'F26' in  f_score_data_dict.keys():datapoint_26   = data_form.datapoint_26  ;   f_score += get_data_score(datapoint_26,   'F26')
                if 'F27' in  f_score_data_dict.keys():datapoint_27   = data_form.datapoint_27  ;   f_score += get_data_score(datapoint_27,   'F27')
                if 'F28' in  f_score_data_dict.keys():datapoint_28   = data_form.datapoint_28  ;   f_score += get_data_score(datapoint_28,   'F28')

                if 'F30' in  f_score_data_dict.keys():datapoint_30   = data_form.datapoint_30  ;   f_score += get_data_score(datapoint_30,   'F30')
                if 'F31' in  f_score_data_dict.keys():datapoint_31   = data_form.datapoint_31  ;   f_score += get_data_score(datapoint_31,   'F31')
                if 'F32' in  f_score_data_dict.keys():datapoint_32   = data_form.datapoint_32  ;   f_score += get_data_score(datapoint_32,   'F32')
                if 'F33' in  f_score_data_dict.keys():datapoint_33   = data_form.datapoint_33  ;   f_score += get_data_score(datapoint_33,   'F33')
                if 'F34' in  f_score_data_dict.keys():datapoint_34   = data_form.datapoint_34  ;   f_score += get_data_score(datapoint_34,   'F34')
                if 'F35' in  f_score_data_dict.keys():datapoint_35   = data_form.datapoint_35  ;   f_score += get_data_score(datapoint_35,   'F35')
                if 'F36' in  f_score_data_dict.keys():datapoint_36   = data_form.datapoint_36  ;   f_score += get_data_score(datapoint_36,   'F36')
                if 'F37' in  f_score_data_dict.keys():datapoint_37   = data_form.datapoint_37  ;   f_score += get_data_score(datapoint_37,   'F37')
                if 'F38' in  f_score_data_dict.keys():datapoint_38   = data_form.datapoint_38  ;   f_score += get_data_score(datapoint_38,   'F38')
                if 'F39' in  f_score_data_dict.keys():datapoint_39   = data_form.datapoint_39  ;   f_score += get_data_score(datapoint_39,   'F39')

                if 'F40' in  f_score_data_dict.keys():datapoint_40   = data_form.datapoint_40  ;   f_score += get_data_score(datapoint_40,   'F40')
                if 'F41' in  f_score_data_dict.keys():datapoint_41   = data_form.datapoint_41  ;   f_score += get_data_score(datapoint_41,   'F41')
                if 'F42' in  f_score_data_dict.keys():datapoint_42   = data_form.datapoint_42  ;   f_score += get_data_score(datapoint_42,   'F42')
                if 'F43' in  f_score_data_dict.keys():datapoint_43   = data_form.datapoint_43  ;   f_score += get_data_score(datapoint_43,   'F43')
                if 'F44' in  f_score_data_dict.keys():datapoint_44   = data_form.datapoint_44  ;   f_score += get_data_score(datapoint_44,   'F44')
                if 'F45' in  f_score_data_dict.keys():datapoint_45   = data_form.datapoint_45  ;   f_score += get_data_score(datapoint_45,   'F45')
                if 'F46' in  f_score_data_dict.keys():datapoint_46   = data_form.datapoint_46  ;   f_score += get_data_score(datapoint_46,   'F46')
                if 'F47' in  f_score_data_dict.keys():datapoint_47   = data_form.datapoint_47  ;   f_score += get_data_score(datapoint_47,   'F47')
                if 'F48' in  f_score_data_dict.keys():datapoint_48   = data_form.datapoint_48  ;   f_score += get_data_score(datapoint_48,   'F48')
                if 'F49' in  f_score_data_dict.keys():datapoint_49   = data_form.datapoint_49  ;   f_score += get_data_score(datapoint_49,   'F49')
  
                if 'F50' in  f_score_data_dict.keys():datapoint_50   = data_form.datapoint_50  ;   f_score += get_data_score(datapoint_50,   'F50')
                if 'F51' in  f_score_data_dict.keys():datapoint_51   = data_form.datapoint_51  ;   f_score += get_data_score(datapoint_51,   'F51')
                if 'F52' in  f_score_data_dict.keys():datapoint_52   = data_form.datapoint_52  ;   f_score += get_data_score(datapoint_52,   'F52')
                if 'F53' in  f_score_data_dict.keys():datapoint_53   = data_form.datapoint_53  ;   f_score += get_data_score(datapoint_53,   'F53')
                if 'F54' in  f_score_data_dict.keys():datapoint_54   = data_form.datapoint_54  ;   f_score += get_data_score(datapoint_54,   'F54')
                if 'F55' in  f_score_data_dict.keys():datapoint_55   = data_form.datapoint_55  ;   f_score += get_data_score(datapoint_55,   'F55')
                if 'F56' in  f_score_data_dict.keys():datapoint_56   = data_form.datapoint_56  ;   f_score += get_data_score(datapoint_56,   'F56')
                if 'F57' in  f_score_data_dict.keys():datapoint_57   = data_form.datapoint_57  ;   f_score += get_data_score(datapoint_57,   'F57')
                if 'F58' in  f_score_data_dict.keys():datapoint_58   = data_form.datapoint_58  ;   f_score += get_data_score(datapoint_58,   'F58')
                if 'F59' in  f_score_data_dict.keys():datapoint_59   = data_form.datapoint_59  ;   f_score += get_data_score(datapoint_59,   'F59')
  
                if 'F60' in  f_score_data_dict.keys():datapoint_60   = data_form.datapoint_60  ;   f_score += get_data_score(datapoint_60,   'F60')
                if 'F61' in  f_score_data_dict.keys():datapoint_61   = data_form.datapoint_61  ;   f_score += get_data_score(datapoint_61,   'F61')
                if 'F62' in  f_score_data_dict.keys():datapoint_62   = data_form.datapoint_62  ;   f_score += get_data_score(datapoint_62,   'F62')
                if 'F63' in  f_score_data_dict.keys():datapoint_63   = data_form.datapoint_63  ;   f_score += get_data_score(datapoint_63,   'F63')
                if 'F64' in  f_score_data_dict.keys():datapoint_64   = data_form.datapoint_64  ;   f_score += get_data_score(datapoint_64,   'F64')
                if 'F65' in  f_score_data_dict.keys():datapoint_65   = data_form.datapoint_65  ;   f_score += get_data_score(datapoint_65,   'F65')
                if 'F66' in  f_score_data_dict.keys():datapoint_66   = data_form.datapoint_66  ;   f_score += get_data_score(datapoint_66,   'F66')
                if 'F67' in  f_score_data_dict.keys():datapoint_67   = data_form.datapoint_67  ;   f_score += get_data_score(datapoint_67,   'F67')
                if 'F68' in  f_score_data_dict.keys():datapoint_68   = data_form.datapoint_68  ;   f_score += get_data_score(datapoint_68,   'F68')
                if 'F69' in  f_score_data_dict.keys():datapoint_69   = data_form.datapoint_69  ;   f_score += get_data_score(datapoint_69,   'F69')
  
                if 'F70' in  f_score_data_dict.keys():datapoint_70   = data_form.datapoint_70  ;   f_score += get_data_score(datapoint_70,   'F70')
                if 'F71' in  f_score_data_dict.keys():datapoint_71   = data_form.datapoint_71  ;   f_score += get_data_score(datapoint_71,   'F71')
                if 'F72' in  f_score_data_dict.keys():datapoint_72   = data_form.datapoint_72  ;   f_score += get_data_score(datapoint_72,   'F72')
                if 'F73' in  f_score_data_dict.keys():datapoint_73   = data_form.datapoint_73  ;   f_score += get_data_score(datapoint_73,   'F73')
                if 'F74' in  f_score_data_dict.keys():datapoint_74   = data_form.datapoint_74  ;   f_score += get_data_score(datapoint_74,   'F74')
                if 'F75' in  f_score_data_dict.keys():datapoint_75   = data_form.datapoint_75  ;   f_score += get_data_score(datapoint_75,   'F75')
                if 'F76' in  f_score_data_dict.keys():datapoint_76   = data_form.datapoint_76  ;   f_score += get_data_score(datapoint_76,   'F76')
                if 'F77' in  f_score_data_dict.keys():datapoint_77   = data_form.datapoint_77  ;   f_score += get_data_score(datapoint_77,   'F77')
                if 'F78' in  f_score_data_dict.keys():datapoint_78   = data_form.datapoint_78  ;   f_score += get_data_score(datapoint_78,   'F78')
                if 'F79' in  f_score_data_dict.keys():datapoint_79   = data_form.datapoint_79  ;   f_score += get_data_score(datapoint_79,   'F79')
  
                if 'F80' in  f_score_data_dict.keys():datapoint_80   = data_form.datapoint_80  ;   f_score += get_data_score(datapoint_80,   'F80')
                if 'F81' in  f_score_data_dict.keys():datapoint_81   = data_form.datapoint_81  ;   f_score += get_data_score(datapoint_81,   'F81')
                if 'F82' in  f_score_data_dict.keys():datapoint_82   = data_form.datapoint_82  ;   f_score += get_data_score(datapoint_82,   'F82')
                if 'F83' in  f_score_data_dict.keys():datapoint_83   = data_form.datapoint_83  ;   f_score += get_data_score(datapoint_83,   'F83')
                if 'F84' in  f_score_data_dict.keys():datapoint_84   = data_form.datapoint_84  ;   f_score += get_data_score(datapoint_84,   'F84')
                if 'F85' in  f_score_data_dict.keys():datapoint_85   = data_form.datapoint_85  ;   f_score += get_data_score(datapoint_85,   'F85')
                if 'F86' in  f_score_data_dict.keys():datapoint_86   = data_form.datapoint_86  ;   f_score += get_data_score(datapoint_86,   'F86')
                if 'F87' in  f_score_data_dict.keys():datapoint_87   = data_form.datapoint_87  ;   f_score += get_data_score(datapoint_87,   'F87')
                if 'F88' in  f_score_data_dict.keys():datapoint_88   = data_form.datapoint_88  ;   f_score += get_data_score(datapoint_88,   'F88')
                if 'F89' in  f_score_data_dict.keys():datapoint_89   = data_form.datapoint_89  ;   f_score += get_data_score(datapoint_89,   'F89')
  
                if 'F90' in  f_score_data_dict.keys():datapoint_90   = data_form.datapoint_90  ;   f_score += get_data_score(datapoint_90,   'F90')
                if 'F91' in  f_score_data_dict.keys():datapoint_91   = data_form.datapoint_91  ;   f_score += get_data_score(datapoint_91,   'F91')
                if 'F92' in  f_score_data_dict.keys():datapoint_92   = data_form.datapoint_92  ;   f_score += get_data_score(datapoint_92,   'F92')
                if 'F93' in  f_score_data_dict.keys():datapoint_93   = data_form.datapoint_93  ;   f_score += get_data_score(datapoint_93,   'F93')
                if 'F94' in  f_score_data_dict.keys():datapoint_94   = data_form.datapoint_94  ;   f_score += get_data_score(datapoint_94,   'F94')
                if 'F95' in  f_score_data_dict.keys():datapoint_95   = data_form.datapoint_95  ;   f_score += get_data_score(datapoint_95,   'F95')
                if 'F96' in  f_score_data_dict.keys():datapoint_96   = data_form.datapoint_96  ;   f_score += get_data_score(datapoint_96,   'F96')
                if 'F97' in  f_score_data_dict.keys():datapoint_97   = data_form.datapoint_97  ;   f_score += get_data_score(datapoint_97,   'F97')
                if 'F98' in  f_score_data_dict.keys():datapoint_98   = data_form.datapoint_98  ;   f_score += get_data_score(datapoint_98,   'F98')
                if 'F99' in  f_score_data_dict.keys():datapoint_99   = data_form.datapoint_99  ;   f_score += get_data_score(datapoint_99,   'F99')

                if 'F100' in  f_score_data_dict.keys():datapoint_100  = data_form.datapoint_100 ;   f_score += get_data_score(datapoint_100,  'F100')
                if 'F101' in  f_score_data_dict.keys():datapoint_101  = data_form.datapoint_101 ;   f_score += get_data_score(datapoint_101,  'F101')
                if 'F102' in  f_score_data_dict.keys():datapoint_102  = data_form.datapoint_102 ;   f_score += get_data_score(datapoint_102,  'F102')
                if 'F103' in  f_score_data_dict.keys():datapoint_103  = data_form.datapoint_103 ;   f_score += get_data_score(datapoint_103,  'F103')
                if 'F104' in  f_score_data_dict.keys():datapoint_104  = data_form.datapoint_104 ;   f_score += get_data_score(datapoint_104,  'F104')
                if 'F105' in  f_score_data_dict.keys():datapoint_105  = data_form.datapoint_105 ;   f_score += get_data_score(datapoint_105,  'F105')
                if 'F106' in  f_score_data_dict.keys():datapoint_106  = data_form.datapoint_106 ;   f_score += get_data_score(datapoint_106,  'F106')
                if 'F107' in  f_score_data_dict.keys():datapoint_107  = data_form.datapoint_107 ;   f_score += get_data_score(datapoint_107,  'F107')
                if 'F108' in  f_score_data_dict.keys():datapoint_108  = data_form.datapoint_108 ;   f_score += get_data_score(datapoint_108,  'F108')
                if 'F109' in  f_score_data_dict.keys():datapoint_109  = data_form.datapoint_109 ;   f_score += get_data_score(datapoint_109,  'F109')

                if 'F110' in  f_score_data_dict.keys():datapoint_110  = data_form.datapoint_110 ;   f_score += get_data_score(datapoint_110,  'F110')
                if 'F111' in  f_score_data_dict.keys():datapoint_111  = data_form.datapoint_111 ;   f_score += get_data_score(datapoint_111,  'F111')
                if 'F112' in  f_score_data_dict.keys():datapoint_112  = data_form.datapoint_112 ;   f_score += get_data_score(datapoint_112,  'F112')
                if 'F113' in  f_score_data_dict.keys():datapoint_113  = data_form.datapoint_113 ;   f_score += get_data_score(datapoint_113,  'F113')
                if 'F114' in  f_score_data_dict.keys():datapoint_114  = data_form.datapoint_114 ;   f_score += get_data_score(datapoint_114,  'F114')
                if 'F115' in  f_score_data_dict.keys():datapoint_115  = data_form.datapoint_115 ;   f_score += get_data_score(datapoint_115,  'F115')
                if 'F116' in  f_score_data_dict.keys():datapoint_116  = data_form.datapoint_116 ;   f_score += get_data_score(datapoint_116,  'F116')
                if 'F117' in  f_score_data_dict.keys():datapoint_117  = data_form.datapoint_117 ;   f_score += get_data_score(datapoint_117,  'F117')
                if 'F118' in  f_score_data_dict.keys():datapoint_118  = data_form.datapoint_118 ;   f_score += get_data_score(datapoint_118,  'F118')
                if 'F119' in  f_score_data_dict.keys():datapoint_119  = data_form.datapoint_119 ;   f_score += get_data_score(datapoint_119,  'F119')



                try:
                    f_score += round(float(str(data_form.profile_data.u_score_p).strip().replace(" ","")),2)
                except:
                    f_score = f_score

                data_form.u_score_b = f_score
                if f_score > 0.5:
                    data_form.u_score = 1.0
                else:
                    data_form.u_score = 0.0
                data_form.save()
            except:
                continue


    del f_score_data_dict

def data_dictionaty_load(request):
    if 'POST' == request.method:
        file_in_memory = request.FILES['file']
        df = pd.read_excel(file_in_memory)

        for index, row in df.iterrows():

            f_code = f'F{str(row["F-code"]).split(" ")[1]}'.strip()
            f_code = f_code if f_code != "" and f_code !="nan" else ""

            f_score = str(row["F-score"]).strip()
            f_score = f_score if f_score != "" and f_score != "nan" else ""


            f_score_a = str(row["F-score A"]).strip()
            f_score_a = f_score_a if f_score_a != "" and f_score_a != "nan" else ""

            f_score_b = str(row["F-score B"]).strip()
            f_score_b = f_score_b if f_score_b != "" and f_score_b != "nan" else ""


            data_point = str(row["Data Point"]).strip()
            data_point = data_point if data_point != "" and data_point != "nan" else ""


            u_score = str(row['U-Score']).strip()
            u_score = u_score if u_score != "" and u_score != "nan" else ""


            link_logic = str(row['Link Logic']).strip()
            link_logic = True if link_logic == "Yes" else False

            data_point_a = str(row['Data point A']).strip()
            data_point_a = data_point_a if data_point_a != "" and data_point_a != "nan" else ""


            data_point_b = str(row['Data point B']).strip()
            data_point_b = data_point_b if data_point_b != "" and data_point_b != "nan" else ""


            data_dictionary_ = str(row['Data Dictionary']).strip()
            data_dictionary_ = data_dictionary_ if data_dictionary_ != "" and data_dictionary_ != "nan" else ""


            display_distenation = str(row['Display Destination']).strip()
            display_distenation = display_distenation if display_distenation != "" and display_distenation != "nan" else ""


            color = str(row['Colour Specs']).split("#")[0].strip()
            color = color if color != "" and color != "nan" else ""




            data_dictionary_form, created = DataDictionary.objects.get_or_create(
                    f_code=f_code,
                    value=data_dictionary_
            )

            if data_dictionary_form:
                print("f_score",f_score)
                data_dictionary_form.f_score = f_score

                data_dictionary_form.f_score_a = f_score_a
                data_dictionary_form.f_score_b = f_score_b
                data_dictionary_form.u_score = u_score
                data_dictionary_form.link_logic = link_logic
                data_dictionary_form.data_point = data_point

                data_dictionary_form.data_point_a = data_point_a
                data_dictionary_form.data_point_b = data_point_b
                data_dictionary_form.display_distenation = display_distenation
                data_dictionary_form.color = color


            data_dictionary_form.save()
        #reload after load all weigh
        print('reload')
        run_f_scores_reloading_weig()



        return redirect('/datadictionary/')

@login_required(login_url='/login/')
def logic_link(request):
    if request.method == "GET":
        search_param = {'is_hide':False}
        page = 1
        data_dictionary = DataDictionary.objects.filter(**search_param).order_by('id')[(page - 1) * 50:50 * page]
        return render(request,'logic_tablwe/logic_table.html',{"data_dictionary":data_dictionary})

@user_is_role_adm
def logic_linkDelete(request,req_id):
    if request.method == "POST":
        pacient = Pacient.objects.get(id=req_id)
        pacient.is_hide = True
        pacient.save()

        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

def logic_linkEdit(request,req_id):
    if request.method == "POST":

        f_code = request.POST.get('edit_f_code')
        f_score = request.POST.get('edit_f_score')

        f_score_a = request.POST.get('edit_f_score_a')

        f_score_b = request.POST.get('edit_f_score_b')

        data_points = request.POST.get('edit_data_points')

        value = request.POST.get('edit_data_dictionary')

        display_destination = request.POST.get('edit_display_destination')

        u_score = request.POST.get('edit_u_score')

        link_logic = request.POST.get('edit_link_logic')
        if link_logic == "Yes":
            link_logic = True
        else:
            link_logic = False



        color_spect = request.POST.get('edit_color_spect')
        data_dictionaty_ = DataDictionary.objects.get(id=req_id)


        data_dictionaty_.color=color_spect
        data_dictionaty_.link_logic=link_logic
        data_dictionaty_.u_score=u_score
        data_dictionaty_.display_distenation=display_destination
        data_dictionaty_.data_point=data_points
        data_dictionaty_.value=value
        data_dictionaty_.f_score_b=f_score_b
        data_dictionaty_.f_score_a=f_score_a
        data_dictionaty_.f_score=f_score
        data_dictionaty_.f_code=f_code


        data_dictionaty_.save()

        return redirect('/datadictionary/')


@login_required(login_url='/login/') #redirect when user is not logged in
def data_dictionaty(request):
    if request.method == "GET":
        search_param = {'is_hide':False}
        page = 1
        data_dictionary = DataDictionary.objects.filter(**search_param).order_by('-id')[(page - 1) * 50:50 * page]
        return render(request,'ddactionary/ddactionary.html',{"data_dictionary":data_dictionary})

    elif request.method=="POST":
        model_code = request.POST.get('model_code')
        f_code = request.POST.get('f_code')
        f_score = request.POST.get('f_score')
        print(f_score)

        f_code = str(model_code)+str(f_code)

        f_score_a = request.POST.get('f_score_a')

        f_score_b = request.POST.get('f_score_b')

        data_points = request.POST.get('data_points')

        value = request.POST.get('data_dictionary')

        display_destination = request.POST.get('display_destination')

        u_score = request.POST.get('u_score')

        link_logic = request.POST.get('link_logic')
        if link_logic == "Yes":
            link_logic = True
        else:
            link_logic = False

        color_spect = request.POST.get('color_spect')

        try:
            print()
            data_dictionary_ =  DataDictionary.objects.get(
                f_code=f_code,
                value=value,
            )
            print('data_dictionary_',data_dictionary_)
            try:

                old_f_score = data_dictionary_.f_score
                print('old_f_score', old_f_score)
                old_f_score = round(float(str(old_f_score).strip().replace(' ','')),2)
                print('old_f_score2',old_f_score)
            except:
                old_f_score = 0.0
        except:

            old_f_score = 0.0
            data_dictionary_= DataDictionary.objects.create(
                f_code=f_code,
                value=value,
                        )
        data_dictionary_.color = color_spect
        data_dictionary_.link_logic = link_logic
        data_dictionary_.u_score = u_score
        data_dictionary_.display_distenation = display_destination
        data_dictionary_.data_point = data_points
        data_dictionary_.f_score_b = f_score_b
        data_dictionary_.f_score_a = f_score_a
        data_dictionary_.f_score = f_score
        data_dictionary_.f_code = f_code

        print('old_f_score2', old_f_score)

        data_dictionary_.save()

        if model_code == 'F' or model_code == 'P':
            print("model_code == 'F' or model_code == 'P'")

            # try:
            f_data_query = {}
            p_data_query = {}

            if "F" in f_code:
                f_data_query['datapoint_'+f_code.replace('F','')] = value

            if "P" in f_code:
                p_data_query['profile_data__datapoint_' + f_code.replace('P','')] = value
            print('f_data_query',f_data_query)
            print('p_data_query',p_data_query)
            data_forms = DataFormTable.objects.filter(
                Q(**f_data_query) | Q(**p_data_query)


            ).distinct()
            print(data_forms)
            # except:
            #     data_forms = []




            for data_form in data_forms:

                try:
                    old_u_score_b = round(float(data_form.u_score_b),2)
                except:
                    old_u_score_b = 0.0



                try:
                    f_score = round(float(f_score),2)
                except:
                    f_score = 0.0

                u_score_b = old_u_score_b - old_f_score + f_score
                data_form.u_score_b =  round(u_score_b,2)
                data_form.save()


        # elif model_code == 'P':
        #     print("model_code == 'P'")
        #     try:
        #         profile_datas = ProfileFormTable.objects.all()
        #
        #     except:
        #         profile_datas = []
        #
        #
        #     for profile_data in profile_datas:
        #
        #
        #         try:
        #             old_u_score_b = round(float(profile_data.u_scu_score_b2ore_p),2)
        #         except:
        #             old_u_score_b = 0.0
        #
        #
        #
        #         try:
        #             f_score = round(float(f_score),2)
        #         except:
        #             f_score = 0.0
        #
        #         u_score_b = float(old_u_score_b) - float(old_f_score) + float(f_score)
        #         profile_data.u_score_p = round(u_score_b,2)
        #         profile_data.save()


        return redirect('/datadictionary/')
def parse_drugs(data):
    return_data = []
    l = re.split(r"(\: \d\,)* unique\:\d\[\]\:", data.replace('\n', ' '))

    for i in l:
        if i:
            prep = i.replace(":", "").replace(",", "").strip()
            if prep and len(prep) >= 3:
                return_data.append(prep)


    if not return_data:
        drugs = Drug.objects.all().values('name')
        for drug_name in drugs:
            if str(drug_name) in str(data):
                return_data.append(drug_name)



    return return_data

def  get_data_score_DB(data,f_code,model_name="Data Form"):
    try:
        data_dict = DataDictionary.objects.get(f_code=f_code,value=data,model_name=model_name)
        value = float(data_dict.f_score)
        if type(value) != float:
            return value
    except:
        return 0.0

def  get_data_score(data,f_code,model_name="Data Form"):
    try:
        data_dict = DataDictionary.objects.get(f_code=f_code,value=data,model_name=model_name)
        value = float(data_dict.f_score)
        if type(value) != float:
            return value
    except:
        return 0.0

@login_required(login_url='/login/')
def master_table_view(request):
    if request.method == "GET":

        search_param = {}

        page = request.GET.get('page') if request.GET.get('page') else 1
        page = int(page)

        first_name = request.GET.get('first_name')

        last_name = request.GET.get('last_name')

        phone = request.GET.get('phone')

        email = request.GET.get('email')

        nsh_id = request.GET.get('nsh_id')

        default_city = request.GET.get('city')

        home_addres = request.GET.get('addres')
        birth_from = request.GET.get('birth_from')

        birth_to = request.GET.get('birth_to')
        search_param['is_hide'] = False
        if email:
            email = vaidator_data.validate_email(str(email))

            search_param['pacient__email'] = email

        if nsh_id:
            nsh_id = str(nsh_id)

            search_param['pacient__nsh_id'] = nsh_id

        if first_name:

            search_param['pacient__first_name'] = first_name

        if last_name:

            search_param['pacient__last_name'] = last_name

        if phone:
            phone = vaidator_data.validate_phone(str(phone))

            search_param['pacient__phone'] = phone

        if default_city:
            default_city = str(default_city)

            search_param['pacient__default_city'] = default_city

        if home_addres:
            home_addres = str(home_addres)

            search_param['pacient__home_addres'] = home_addres


        master_table = DataFormTable.objects.filter(**search_param).order_by('create_datetime')[(page-1)*50:50*page]
        return render(request,'masterTable/masterTable.html',{
            'master_table':master_table,'user':request.user,
        })

    if request.method == "POST":
        f_score_data_dict = create_f_score_data_dict()

        def get_data_score(value, f_code):

            try:
                if str(value) != "" and str(value) != None and f_code in f_score_data_dict.keys():
                    print('value,f_code:', value, f_code)
                    try:
                        fcode_val = f_score_data_dict[str(f_code)][str(value).lower()]
                        print('fcode_val', fcode_val)
                        fcode_val = round(float(str(fcode_val).strip().replace(" ", '')), 2)
                        print('fcode_val', fcode_val)
                        return fcode_val
                    except:
                        return 0.0

                else:
                    return 0.0

            except:
                return 0.0
        try:
            file_ProfileForm = request.FILES['file_ProfileForm']
            file_ProfileForm = file_ProfileForm if file_ProfileForm else None
        except:
            file_ProfileForm = None


        try:

            file_DataForm = request.FILES['file_DataForm']
            file_DataForm = file_DataForm if file_DataForm else None

            dfs = pd.read_excel(file_DataForm)
            # from pyexcel_xls import get_data as xls_get
            # from pyexcel_xlsx import get_data as xlsx_get

            # file_DataForm = request.FILES['file_DataForm']
            # pd.read_excel(file_DataForm, )
            # print("file_DataForm",file_DataForm)
            # if (str(file_DataForm).split('.')[-1] == 'xls'):
            #     data = xls_get(file_DataForm, column_limit=120)
            #     print("data", data)
            # elif (str(file_DataForm).split('.')[-1] == 'xlsx'):
            #     data = xlsx_get(file_DataForm, column_limit=120)
            #     print("data", data)
            for i in dfs:
                print(i)
        except:
            file_DataForm = None

        google_drive_url_ProfileForm = request.POST.get('google_drive_url_ProfileForm')
        google_drive_url_ProfileForm = google_drive_url_ProfileForm if google_drive_url_ProfileForm else None

        google_drive_url_DataForm = request.POST.get('google_drive_url_DataForm')
        google_drive_url_DataForm = google_drive_url_DataForm if google_drive_url_DataForm else None

        if google_drive_url_DataForm:

            gg_drive_proccess1 = GoogleDrive(google_drive_url_DataForm)
            gg_drive_proccess1.request()
            gg_drive_proccess1.gog_drive_to_json()
            gg_drive_proccess1.data_from_jjson()
            for row_ in gg_drive_proccess1.table_data:
                pacient_id = procces_pacient(row_[:8])


                if pacient_id ==  None: continue

                create_datetime = vaidator_data.validate_datetime(row_[0])

                try:
                    profile_form = ProfileFormTable.objects.filter(
                        pacient_id=pacient_id,
                    ).order_by("-create_datetime")[0]
                except:
                    profile_form = ProfileFormTable.objects.create(pacient_id=pacient_id)
                    profile_form.save()

                try:
                    nsh_id_gdisk = row_[120]

                    data_form, created = DataFormTable.objects.get_or_create(
                        nsh_id_gdisk=nsh_id_gdisk
                    )
                except:
                    data_form, created = DataFormTable.objects.get_or_create(
                        pacient_id=pacient_id,
                        create_datetime=create_datetime,
                        profile_data_id=profile_form.id
                    )

                if data_form:
                    f_score = 0

                    data_form.datapoint_7  = row_[6]       ; f_score += get_data_score(row_[6], 'F7')
                    data_form.datapoint_8  = row_[7]       ; f_score += get_data_score(row_[7], 'F8')
                    data_form.datapoint_9  = row_[8]       ; f_score += get_data_score(row_[8], 'F9')
                    data_form.datapoint_10 = row_[9]       ; f_score += get_data_score(row_[9], 'F10')


                    data_form.datapoint_11 = row_[10]     ; f_score += get_data_score(row_[10], 'F11')
                    data_form.datapoint_12 = row_[11]     ; f_score += get_data_score(row_[11], 'F12')
                    data_form.datapoint_13 = row_[12]     ; f_score += get_data_score(row_[12], 'F13')
                    data_form.datapoint_14 = row_[13]     ; f_score += get_data_score(row_[13], 'F14')
                    data_form.datapoint_15 = row_[14]     ; f_score += get_data_score(row_[14], 'F15')
                    data_form.datapoint_16 = row_[15]     ; f_score += get_data_score(row_[15], 'F16')

                    data_form.datapoint_16a = row_[16]   ; f_score += get_data_score(row_[16], 'F16_a')
                    data_form.datapoint_16b = row_[17]   ; f_score += get_data_score(row_[17], 'F16_b')
                    data_form.datapoint_17 = row_[18]    ; f_score += get_data_score(row_[18], 'F17')
                    data_form.datapoint_18 = row_[19]    ; f_score += get_data_score(row_[19], 'F18')
                    data_form.datapoint_19 = row_[20]    ; f_score += get_data_score(row_[20], 'F19')


                    data_form.datapoint_20 = row_[21]    ;f_score += get_data_score(row_[21], 'F20')

                    data_form.datapoint_21 = row_[22]    ;f_score += get_data_score(row_[22], 'F21')
                    data_form.datapoint_22 = row_[23]    ;f_score += get_data_score(row_[23], 'F22')
                    data_form.datapoint_23 = row_[24]    ;f_score += get_data_score(row_[24], 'F23')
                    data_form.datapoint_24 = row_[25]    ;f_score += get_data_score(row_[25], 'F24')
                    data_form.datapoint_25 = row_[26]    ;f_score += get_data_score(row_[26], 'F25')
                    data_form.datapoint_26 = row_[27]    ;f_score += get_data_score(row_[27], 'F26')
                    data_form.datapoint_27 = row_[28]    ;f_score += get_data_score(row_[28], 'F27')
                    data_form.datapoint_28 = row_[29]    ;f_score += get_data_score(row_[29], 'F28')


                    data_form.datapoint_30 = row_[31]    ;f_score += get_data_score(row_[31], 'F30')
                    data_form.datapoint_31 = row_[32]    ;f_score += get_data_score(row_[32], 'F31')
                    data_form.datapoint_32 = row_[33]    ;f_score += get_data_score(row_[33], 'F32')
                    data_form.datapoint_33 = row_[34]    ;f_score += get_data_score(row_[34], 'F33')
                    data_form.datapoint_34 = row_[35]    ;f_score += get_data_score(row_[35], 'F34')
                    data_form.datapoint_35 = row_[36]    ;f_score += get_data_score(row_[36], 'F35')
                    data_form.datapoint_36 = row_[37]    ;f_score += get_data_score(row_[37], 'F36')
                    data_form.datapoint_37 = row_[38]    ;f_score += get_data_score(row_[38], 'F37')
                    data_form.datapoint_38 = row_[39]    ;f_score += get_data_score(row_[39], 'F38')
                    data_form.datapoint_39 = row_[40]    ;f_score += get_data_score(row_[40], 'F39')


                    data_form.datapoint_40 = row_[41]    ;f_score += get_data_score(row_[41], 'F40')
                    data_form.datapoint_41 = row_[42]    ;f_score += get_data_score(row_[42], 'F41')
                    data_form.datapoint_42 = row_[43]    ;f_score += get_data_score(row_[43], 'F42')
                    data_form.datapoint_43 = row_[44]    ;f_score += get_data_score(row_[44], 'F43')
                    data_form.datapoint_44 = row_[45]    ;f_score += get_data_score(row_[45], 'F44')
                    data_form.datapoint_45 = row_[46]    ;f_score += get_data_score(row_[46], 'F45')
                    data_form.datapoint_46 = row_[47]    ;f_score += get_data_score(row_[47], 'F46')
                    data_form.datapoint_47 = row_[48]    ;f_score += get_data_score(row_[48], 'F47')
                    data_form.datapoint_48 = row_[49]    ;f_score += get_data_score(row_[49], 'F48')
                    data_form.datapoint_49 = row_[50]    ;f_score += get_data_score(row_[50], 'F49')


                    data_form.datapoint_50 = row_[51]    ;f_score += get_data_score(row_[51], 'F50')
                    data_form.datapoint_51 = row_[52]    ;f_score += get_data_score(row_[52], 'F51')
                    data_form.datapoint_52 = row_[53]    ;f_score += get_data_score(row_[53], 'F52')
                    data_form.datapoint_53 = row_[54]    ;f_score += get_data_score(row_[54], 'F53')
                    data_form.datapoint_54 = row_[55]    ;f_score += get_data_score(row_[55], 'F54')
                    data_form.datapoint_55 = row_[56]    ;f_score += get_data_score(row_[56], 'F55')
                    data_form.datapoint_56 = row_[57]    ;f_score += get_data_score(row_[57], 'F56')
                    data_form.datapoint_57 = row_[58]    ;f_score += get_data_score(row_[58], 'F57')
                    data_form.datapoint_58 = row_[59]    ;f_score += get_data_score(row_[59], 'F58')
                    data_form.datapoint_59 = row_[60]    ;f_score += get_data_score(row_[60], 'F59')


                    data_form.datapoint_60 = row_[61]   ;f_score += get_data_score(row_[61], 'F60')
                    data_form.datapoint_61 = row_[62]   ;f_score += get_data_score(row_[62], 'F61')
                    data_form.datapoint_62 = row_[63]   ;f_score += get_data_score(row_[63], 'F62')
                    data_form.datapoint_63 = row_[64]   ;f_score += get_data_score(row_[64], 'F63')
                    data_form.datapoint_64 = row_[65]   ;f_score += get_data_score(row_[65], 'F64')
                    data_form.datapoint_65 = row_[66]   ;f_score += get_data_score(row_[66], 'F65')
                    data_form.datapoint_66 = row_[67]   ;f_score += get_data_score(row_[67], 'F66')
                    data_form.datapoint_67 = row_[68]   ;f_score += get_data_score(row_[68], 'F67')
                    data_form.datapoint_68 = row_[69]   ;f_score += get_data_score(row_[69], 'F68')
                    data_form.datapoint_69 = row_[70]   ;f_score += get_data_score(row_[70], 'F69')




                    data_form.datapoint_70 = row_[71]   ;f_score += get_data_score(row_[71], 'F70')
                    data_form.datapoint_71 = row_[72]   ;f_score += get_data_score(row_[72], 'F71')
                    data_form.datapoint_72 = row_[73]   ;f_score += get_data_score(row_[73], 'F72')
                    data_form.datapoint_73 = row_[74]   ;f_score += get_data_score(row_[74], 'F73')
                    data_form.datapoint_74 = row_[75]   ;f_score += get_data_score(row_[75], 'F74')
                    data_form.datapoint_75 = row_[76]   ;f_score += get_data_score(row_[76], 'F75')
                    data_form.datapoint_76 = row_[77]   ;f_score += get_data_score(row_[77], 'F76')
                    data_form.datapoint_77 = row_[78]   ;f_score += get_data_score(row_[78], 'F77')
                    data_form.datapoint_78 = row_[79]   ;f_score += get_data_score(row_[79], 'F78')
                    data_form.datapoint_79 = row_[80]   ;f_score += get_data_score(row_[80], 'F79')


                    data_form.datapoint_80 = row_[81]   ;f_score += get_data_score(row_[81], 'F80')
                    data_form.datapoint_81 = row_[82]   ;f_score += get_data_score(row_[82], 'F81')
                    data_form.datapoint_82 = row_[83]   ;f_score += get_data_score(row_[83], 'F82')
                    data_form.datapoint_83 = row_[84]   ;f_score += get_data_score(row_[84], 'F83')
                    data_form.datapoint_84 = row_[85]   ;f_score += get_data_score(row_[85], 'F84')
                    data_form.datapoint_85 = row_[86]   ;f_score += get_data_score(row_[86], 'F85')
                    data_form.datapoint_86 = row_[87]   ;f_score += get_data_score(row_[87], 'F86')
                    data_form.datapoint_87 = row_[88]   ;f_score += get_data_score(row_[88], 'F87')
                    data_form.datapoint_88 = row_[89]   ;f_score += get_data_score(row_[89], 'F88')
                    data_form.datapoint_89 = row_[90]   ;f_score += get_data_score(row_[90], 'F89')


                    data_form.datapoint_90 = row_[91]    ;f_score += get_data_score(row_[91], 'F90')
                    data_form.datapoint_91 = row_[92]    ;f_score += get_data_score(row_[92], 'F91')
                    data_form.datapoint_92 = row_[93]    ;f_score += get_data_score(row_[93], 'F92')
                    data_form.datapoint_93 = row_[94]    ;f_score += get_data_score(row_[94], 'F93')
                    data_form.datapoint_94 = row_[95]    ;f_score += get_data_score(row_[95], 'F94')
                    data_form.datapoint_95 = row_[96]    ;f_score += get_data_score(row_[96], 'F95')
                    data_form.datapoint_96 = row_[97]    ;f_score += get_data_score(row_[97], 'F96')
                    data_form.datapoint_97 = row_[98]    ;f_score += get_data_score(row_[98], 'F97')
                    data_form.datapoint_98 = row_[99]    ;f_score += get_data_score(row_[99], 'F98')
                    data_form.datapoint_99 = row_[100]   ;f_score += get_data_score(row_[100], 'F99')

                    data_form.datapoint_100 = row_[101]  ;f_score += get_data_score(row_[101], 'F100')
                    data_form.datapoint_101 = row_[102]  ;f_score += get_data_score(row_[102], 'F101')
                    data_form.datapoint_102 = row_[103]  ;f_score += get_data_score(row_[103], 'F102')
                    data_form.datapoint_103 = row_[104]  ;f_score += get_data_score(row_[104], 'F103')
                    data_form.datapoint_104 = row_[105]  ;f_score += get_data_score(row_[105], 'F104')
                    data_form.datapoint_105 = row_[106]  ;f_score += get_data_score(row_[106], 'F105')
                    data_form.datapoint_106 = row_[107]  ;f_score += get_data_score(row_[107], 'F106')
                    data_form.datapoint_107 = row_[108]  ;f_score += get_data_score(row_[108], 'F107')
                    data_form.datapoint_108 = row_[109]  ;f_score += get_data_score(row_[109], 'F108')
                    data_form.datapoint_109 = row_[110]  ;f_score += get_data_score(row_[110], 'F109')

                    data_form.datapoint_110 = row_[111]  ;f_score += get_data_score(row_[111], 'F110')
                    data_form.datapoint_111 = row_[112]  ;f_score += get_data_score(row_[112], 'F111')
                    data_form.datapoint_112 = row_[113]  ;f_score += get_data_score(row_[113], 'F112')
                    data_form.datapoint_113 = row_[114]  ;f_score += get_data_score(row_[114], 'F113')
                    data_form.datapoint_114 = row_[115]  ;f_score += get_data_score(row_[115], 'F114')
                    data_form.datapoint_115 = row_[116]  ;f_score += get_data_score(row_[116], 'F115')
                    data_form.datapoint_116 = row_[117]  ;f_score += get_data_score(row_[117], 'F116')
                    data_form.datapoint_117 = row_[118]  ;f_score += get_data_score(row_[118], 'F117')
                    data_form.datapoint_118 = row_[119]  ;f_score += get_data_score(row_[119], 'F118')
                    data_form.datapoint_119 = row_[120]  ;f_score += get_data_score(row_[119], 'F119')


                    data_form.datapoint_120 = row_[121]


                    data_form.u_score_b = f_score

                    data_form.u_score = profile_form.u_score_p + data_form.u_score_b





                    data_form.save()



        if google_drive_url_ProfileForm:

            gg_drive_proccess2 = GoogleDrive(google_drive_url_ProfileForm)
            gg_drive_proccess2.request()
            gg_drive_proccess2.gog_drive_to_json()
            gg_drive_proccess2.data_from_jjson()

            for row_ in gg_drive_proccess2.table_data:
                pacient_id = procces_pacient(row_[:8])
                if pacient_id == None: continue
                else:

                    create_datetime = vaidator_data.validate_datetime(row_[0])


                    try:
                        nsh_id_gdisk = row_[96]
                        profile_form, created = ProfileFormTable.objects.get_or_create(
                            nsh_id_gdisk=nsh_id_gdisk
                        )
                    except:
                        profile_form, created = ProfileFormTable.objects.get_or_create(
                            pacient_id=pacient_id,
                            create_datetime=create_datetime)


                    if profile_form:
                        f_score = 0

                        profile_form.sex = row_[8]              ; f_score += get_data_score(row_[8], 'P9')
                        profile_form.datapoint_10 = row_[9]     ; f_score += get_data_score(row_[9], 'P10')

                        metic_sys = row_[10]                    ; f_score += get_data_score(row_[9], 'P11')
                        profile_form.metic_sys = metic_sys

                        profile_form.datapoint_12 = row_[11]    ; f_score += get_data_score(row_[11], 'P12')
                        profile_form.datapoint_13a = row_[12]   ; f_score += get_data_score(row_[12], 'P13_a')
                        profile_form.datapoint_13b = row_[13]   ; f_score += get_data_score(row_[13], 'P13_b')
                        profile_form.datapoint_14 = row_[14]    ; f_score += get_data_score(row_[14], 'P14')
                        profile_form.datapoint_15 = row_[15]    ; f_score += get_data_score(row_[15], 'P15')
                        profile_form.datapoint_16 = row_[16]    ; f_score += get_data_score(row_[16], 'P16')
                        profile_form.datapoint_17 = row_[17]    ; f_score += get_data_score(row_[17], 'P17')
                        profile_form.datapoint_18 = row_[18]    ; f_score += get_data_score(row_[18], 'P18')
                        profile_form.datapoint_19 = row_[19]    ; f_score += get_data_score(row_[19], 'P19')

                        profile_form.datapoint_20 = row_[20]    ; f_score += get_data_score(row_[20], 'P20')
                        profile_form.datapoint_21 = row_[21]    ; f_score += get_data_score(row_[21], 'P21')
                        profile_form.datapoint_22 = row_[22]    ; f_score += get_data_score(row_[22], 'P22')
                        profile_form.datapoint_23 = row_[23]    ; f_score += get_data_score(row_[23], 'P23')
                        profile_form.datapoint_24 = row_[24]    ; f_score += get_data_score(row_[24], 'P24')
                        profile_form.datapoint_25 = row_[25]    ; f_score += get_data_score(row_[25], 'P25')
                        profile_form.datapoint_26 = row_[26]    ; f_score += get_data_score(row_[26], 'P26')
                        profile_form.datapoint_27 = row_[27]    ; f_score += get_data_score(row_[27], 'P27')
                        profile_form.datapoint_28 = row_[28]    ; f_score += get_data_score(row_[28], 'P28')

                        datapoint_29 = row_[29]                 ; f_score += get_data_score(row_[29], 'P29')
                        if datapoint_29:
                            datapoint_29_list_id = []
                            datapoint_29_list = parse_drugs(datapoint_29)
                            if datapoint_29_list:
                                for datapoint_29_name in datapoint_29_list:
                                    drug,created_dr1 = Drug.objects.get_or_create(name=datapoint_29_name)
                                    if created_dr1:
                                        drug.save()
                                    datapoint_29_list_id.append(drug)
                            profile_form.datapoint_29.add(*datapoint_29_list_id)


                        datapoint_30 = row_[30]                 ; f_score += get_data_score(row_[30], 'P30')
                        if datapoint_30:
                            datapoint_30_list_id = []
                            datapoint_30_list = parse_drugs(datapoint_30)
                            if datapoint_30_list:
                                for datapoint_30_name in datapoint_30_list:
                                    drug1, created_dr2 = Drug.objects.get_or_create(name=datapoint_30_name)
                                    if created_dr2:
                                        drug1.save()
                                    datapoint_30_list_id.append(drug1)

                            profile_form.datapoint_30.add(*datapoint_30_list_id)

                        datapoint_31 = row_[31]                 ; f_score += get_data_score(row_[31], 'P31')

                        if datapoint_31:
                            datapoint_31_list_id = []
                            datapoint_31_list = parse_operation(datapoint_31)
                            if datapoint_31_list:
                                for n,datapoint_31_vars in enumerate(datapoint_31_list):
                                    operation, created_op = MedicalOperation.objects.get_or_create(
                                        title=datapoint_31_vars[1],
                                        date_operation=vaidator_data.validate_date(datapoint_31_vars[2]),
                                        num_operation=datapoint_31_vars[0]
                                        )
                                    if created_op:
                                        operation.save()
                                    datapoint_31_list_id.append(operation)

                            profile_form.datapoint_31.add(*datapoint_31_list_id)


                        profile_form.datapoint_32 = row_[32]   ; f_score += get_data_score(row_[32], 'P32')

                        datapoint_33 = row_[33]
                        datapoint_33 = vaidator_data.validate_date(datapoint_33) if vaidator_data.validate_date(datapoint_33) else None
                        profile_form.datapoint_33 = datapoint_33  ;f_score += get_data_score(row_[33], 'P33')
                        profile_form.datapoint_34 = row_[34]  ; f_score += get_data_score(row_[34], 'P34')
                        profile_form.datapoint_35 = row_[35]  ; f_score += get_data_score(row_[35], 'P35')
                        profile_form.datapoint_36 = row_[36]  ; f_score += get_data_score(row_[36], 'P36')
                        profile_form.datapoint_37 = row_[37]  ; f_score += get_data_score(row_[37], 'P37')
                        profile_form.datapoint_38 = row_[38]  ; f_score += get_data_score(row_[38], 'P38')
                        profile_form.datapoint_39 = row_[39]  ; f_score += get_data_score(row_[39], 'P39')

                        profile_form.datapoint_40 = row_[40] ;  f_score += get_data_score(row_[40], 'P40')
                        profile_form.datapoint_41 = row_[41] ;  f_score += get_data_score(row_[41], 'P41')
                        profile_form.datapoint_42 = row_[42] ;  f_score += get_data_score(row_[42], 'P42')
                        profile_form.datapoint_43 = row_[43] ;  f_score += get_data_score(row_[43], 'P43')
                        profile_form.datapoint_44 = row_[44] ;  f_score += get_data_score(row_[44], 'P44')
                        profile_form.datapoint_45 = row_[45] ;  f_score += get_data_score(row_[45], 'P45')
                        profile_form.datapoint_46 = row_[46] ;  f_score += get_data_score(row_[46], 'P46')
                        profile_form.datapoint_47 = row_[47] ;  f_score += get_data_score(row_[47], 'P47')
                        profile_form.datapoint_48 = row_[48] ;  f_score += get_data_score(row_[48], 'P48')
                        profile_form.datapoint_49 = row_[49] ;  f_score += get_data_score(row_[49], 'P49')

                        profile_form.datapoint_50 = row_[50]    ; f_score += get_data_score(row_[50], 'P50')
                        profile_form.datapoint_51 = row_[51]    ; f_score += get_data_score(row_[51], 'P51')

                        datapoint_52 = row_[52]                 ; f_score += get_data_score(row_[52], 'P52')

                        if datapoint_52:
                            datapoint_52_list_id = []
                            datapoint_52_list = parse_datapoint_52(datapoint_52)
                            if datapoint_52_list:
                                for n,datapoint_52_vars in enumerate(datapoint_52_list):
                                    chldPacientDurBreast, created_chld = ChldPacientDurBreast.objects.get_or_create(
                                        num_child=int(datapoint_52_vars[0]),
                                        period_months=int(datapoint_52_vars[1]),
                                        )
                                    if created_chld:
                                        chldPacientDurBreast.save()
                                    datapoint_52_list_id.append(chldPacientDurBreast)

                            profile_form.datapoint_52.add(*datapoint_52_list_id)

                        profile_form.datapoint_53 = row_[53]    ; f_score += get_data_score(row_[53], 'P53')
                        profile_form.datapoint_54 = row_[54]    ; f_score += get_data_score(row_[54], 'P54')
                        profile_form.datapoint_55 = row_[55]    ; f_score += get_data_score(row_[55], 'P55')
                        profile_form.datapoint_56 = row_[56]    ; f_score += get_data_score(row_[56], 'P56')
                        profile_form.datapoint_57 = row_[57]    ; f_score += get_data_score(row_[57], 'P57')
                        profile_form.datapoint_58 = row_[58]    ; f_score += get_data_score(row_[58], 'P58')
                        profile_form.datapoint_59 = row_[59]    ; f_score += get_data_score(row_[59], 'P59')


                        profile_form.datapoint_60 = row_[60]    ; f_score += get_data_score(row_[60], 'P60')
                        profile_form.datapoint_61 = row_[61]    ; f_score += get_data_score(row_[61], 'P61')

                        datapoint_62 = row_[62]                 ; f_score += get_data_score(row_[62], 'P62')
                        if datapoint_62:
                            datapoint_62_list_id = []
                            datapoint_62_list = parse_datapoint_62(datapoint_62,4)
                            if datapoint_62_list:
                                for n,datapoint_62_vars in enumerate(datapoint_62_list):
                                    cancer_,cancer_new = Cancer.objects.get_or_create(title=datapoint_62_vars[1])
                                    chldPacientDurBreast, created_chld = RelationhipPacient.objects.get_or_create(
                                        type_cancer=cancer_,
                                        relation=datapoint_62_vars[0],
                                        age_diagnise=int(datapoint_62_vars[2]),
                                        position_cancer=datapoint_62_vars[3]
                                        )
                                    if created_chld:
                                        chldPacientDurBreast.save()
                                    datapoint_62_list_id.append(chldPacientDurBreast)

                            profile_form.datapoint_62.add(*datapoint_62_list_id)



                        profile_form.datapoint_63 = row_[63]    ; f_score += get_data_score(row_[63], 'P63')

                        datapoint_64 = row_[64]                 ; f_score += get_data_score(row_[64], 'P64')
                        if datapoint_64:
                            datapoint_64_list_id = []
                            datapoint_64_list = parse_datapoint_62(datapoint_64,3)
                            if datapoint_64_list:
                                for n, datapoint_64_vars in enumerate(datapoint_64_list):

                                    cancer,cr_cancer = Cancer.objects.get_or_create(title=datapoint_64_vars[1])
                                    chldPacientDurBreast, created_chld = RelationhipPacient.objects.get_or_create(
                                        type_cancer=cancer,
                                        relation=datapoint_64_vars[0],
                                        age_diagnise=int(datapoint_64_vars[2]),
                                    )
                                    if created_chld:
                                        chldPacientDurBreast.save()
                                    datapoint_64_list_id.append(chldPacientDurBreast)

                            profile_form.datapoint_64.add(*datapoint_64_list_id)


                        profile_form.datapoint_65 = row_[65]    ; f_score += get_data_score(row_[65], 'P65')





                        datapoint_66 = row_[66]                 ; f_score += get_data_score(row_[66], 'P66')

                        if datapoint_66:
                            datapoint_66_list_id = []
                            datapoint_66_list = parse_datapoint_62(datapoint_66)
                            if datapoint_66_list:
                                for n, datapoint_66_vars in enumerate(datapoint_66_list):

                                    cancer,create_cancer = Cancer.objects.get_or_create(title=datapoint_66_vars[1])
                                    cancer_h, create_cancer_h = CancerHistory.objects.get_or_create(
                                        cancer=cancer,
                                        date_detect=vaidator_data.validate_date(datapoint_66_vars[2])
                                    )

                                    if create_cancer_h:
                                        cancer_h.save()
                                    datapoint_66_list_id.append(cancer_h)

                            profile_form.datapoint_66.add(*datapoint_66_list_id)


                        profile_form.datapoint_67 = row_[67]    ; f_score += get_data_score(row_[67], 'P67')
                        profile_form.datapoint_68 = row_[68]    ; f_score += get_data_score(row_[68], 'P68')
                        profile_form.datapoint_69 = row_[69]    ; f_score += get_data_score(row_[69], 'P69')

                        profile_form.datapoint_70 = row_[70]    ; f_score += get_data_score(row_[70], 'P70')
                        profile_form.datapoint_71 = row_[71]    ; f_score += get_data_score(row_[71], 'P71')
                        profile_form.datapoint_72 = row_[72]    ; f_score += get_data_score(row_[72], 'P72')
                        profile_form.datapoint_73 = row_[73]    ; f_score += get_data_score(row_[73], 'P73')
                        profile_form.datapoint_74 = row_[74]    ; f_score += get_data_score(row_[74], 'P74')
                        profile_form.datapoint_75 = row_[75]    ; f_score += get_data_score(row_[75], 'P75')
                        profile_form.datapoint_76 = row_[76]    ; f_score += get_data_score(row_[76], 'P76')
                        profile_form.datapoint_77 = row_[77]    ; f_score += get_data_score(row_[77], 'P77')
                        profile_form.datapoint_78 = row_[78]    ; f_score += get_data_score(row_[78], 'P78')
                        profile_form.datapoint_79 = row_[79]    ; f_score += get_data_score(row_[79], 'P79')

                        profile_form.datapoint_80 = row_[80]    ; f_score += get_data_score(row_[80], 'P80')
                        profile_form.datapoint_81 = row_[81]    ; f_score += get_data_score(row_[81], 'P81')
                        profile_form.datapoint_82 = row_[82]    ; f_score += get_data_score(row_[82], 'P82')
                        profile_form.datapoint_83 = row_[83]    ; f_score += get_data_score(row_[83], 'P83')
                        profile_form.datapoint_84 = row_[84]    ; f_score += get_data_score(row_[84], 'P84')
                        profile_form.datapoint_85 = row_[85]    ; f_score += get_data_score(row_[85], 'P85')
                        profile_form.datapoint_86 = row_[86]    ; f_score += get_data_score(row_[86], 'P86')
                        profile_form.datapoint_87 = row_[87]    ; f_score += get_data_score(row_[87], 'P87')
                        profile_form.datapoint_88 = row_[88]    ; f_score += get_data_score(row_[88], 'P88')
                        profile_form.datapoint_89 = row_[89]    ; f_score += get_data_score(row_[89], 'P89')

                        profile_form.datapoint_90 = row_[90]    ; f_score += get_data_score(row_[90], 'P90')
                        profile_form.datapoint_91 = row_[91]    ; f_score += get_data_score(row_[91], 'P91')
                        profile_form.datapoint_92 = row_[92]    ; f_score += get_data_score(row_[92], 'P92')
                        profile_form.datapoint_93 = row_[93]    ; f_score += get_data_score(row_[93], 'P93')
                        profile_form.datapoint_94 = row_[94]    ; f_score += get_data_score(row_[94], 'P94')
                        profile_form.datapoint_95 = row_[95]    ; f_score += get_data_score(row_[95], 'P95')


                        try:
                            profile_form.nsh_id_gdisk = row_[96]
                        except:
                            print("err", row_[93:])
                        profile_form.u_score_p = f_score
                    profile_form.save()
                    try:
                        data_form = DataFormTable.objects.filter(
                            pacient_id=pacient_id,
                            profile_data=profile_form,
                        ).order_by('-create_datetime')[0]
                    except:
                        data_form = DataFormTable.objects.create(
                            pacient_id=pacient_id,
                            profile_data=profile_form,
                            create_datetime=create_datetime
                        )



                    data_form.u_score = profile_form.u_score_p + data_form.u_score_b
                    data_form.save()



        del f_score_data_dict


        return render(request, 'desctop/succsespage/suucsespage.html')



def pacientDelete(request,req_id):
    if request.method == "POST":
        pacient = Pacient.objects.get(id=req_id)
        pacient.is_hide = True
        pacient.save()

        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

def masterTableDelete(request,req_id):
    if request.method == "POST":
        mastab = DataFormTable.objects.get(id=req_id)
        mastab.is_hide = True
        mastab.save()

        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

def pacientDownloadFile(request):
    if request.method =='GET':
        format_file = request.GET.get('format_file')

        format_file = format_file if format_file else 'csv'

        search_param = {}

        first_name = request.GET.get('first_name')

        last_name = request.GET.get('last_name')

        phone = request.GET.get('phone')

        email = request.GET.get('email')

        nsh_id = request.GET.get('nsh_id')

        default_city = request.GET.get('city')

        home_addres = request.GET.get('addres')
        birth_from = request.GET.get('birth_from')

        birth_to = request.GET.get('birth_to')

        if email:
            email = vaidator_data.validate_email(str(email))

            search_param['email'] = email

        if nsh_id:
            nsh_id = str(nsh_id)

            search_param['nsh_id'] = nsh_id

        if first_name:

            search_param['first_name'] = first_name

        if last_name:

            search_param['last_name'] = last_name

        if phone:
            phone = vaidator_data.validate_phone(str(phone))

            search_param['phone'] = phone

        if default_city:
            default_city = str(default_city)

            search_param['default_city'] = default_city

        if home_addres:
            home_addres = str(home_addres)

            search_param['home_addres'] = home_addres
        search_param['is_hide'] = False
        # if birth_from and birth_to:
        #     birth_from = vaidator_data.validate_date(str(birth_from))
        #
        #     search_param['date_of_birth__range'] = [birth_from, birth_to]
        #
        # elif birth_from:
        #     birth_from = vaidator_data.validate_date(str(birth_from))
        #
        #     search_param['date_of_birth__range'] = [birth_from, '2023-12-30']
        #
        # elif birth_to:
        #     birth_to = vaidator_data.validate_date(str(birth_to))
        #
        #     search_param['date_of_birth__range'] = ['1900-01-01',birth_to]
        #
        # else:
        #     search_param['date_of_birth__range'] = ['1900-01-01','2023-12-30']

        pacients = Pacient.objects.filter(
            **search_param
        ).order_by('-id')

        columns_names = ['ID', 'NSH ID', 'First Name', 'Last Name', 'Date of birth',
                         'Pacient phone', 'Email', 'City', 'Address']

        if format_file == 'csv':



            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="clinical_pacients.csv"'

            writer = csv.writer(response)
            writer.writerow(columns_names)

            for pacient_data in pacients:
                writer.writerow([
                    pacient_data.id,
                    pacient_data.nsh_id,
                    pacient_data.first_name,
                    pacient_data.last_name,
                    str(pacient_data.date_of_birth),
                    pacient_data.phone,
                    pacient_data.email,
                    pacient_data.default_city,
                    pacient_data.home_addres,


                                 ])

            return response

        elif format_file == 'xls':

            # content-type of response
            response = HttpResponse(content_type='application/ms-excel')

            # decide file name
            response['Content-Disposition'] = 'attachment; filename="clicin_pacient.xls"'

            # creating workbook
            wb = xlwt.Workbook(encoding='utf-8')

            # adding sheet
            ws = wb.add_sheet("sheet1")

            # Sheet header, first row
            row_num = 0

            font_style = xlwt.XFStyle()
            # headers are bold
            font_style.font.bold = True

            # column header names, you can use your own headers here
            columns = columns_names

            # write column headers in sheet
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num], font_style)

            # Sheet body, remaining rows
            font_style = xlwt.XFStyle()

            # get your data, from database or from a text file...
            for pacient in pacients:
                row_num = row_num + 1

                ws.write(row_num, 0, pacient.id, font_style)
                ws.write(row_num, 1, pacient.nsh_id, font_style)
                ws.write(row_num, 2, pacient.first_name, font_style)
                ws.write(row_num, 3, pacient.last_name, font_style)
                ws.write(row_num, 4, str(pacient.date_of_birth), font_style)
                ws.write(row_num, 5, pacient.phone, font_style)
                ws.write(row_num, 6, pacient.email, font_style)
                ws.write(row_num, 7, pacient.default_city, font_style)
                ws.write(row_num, 8, pacient.home_addres, font_style)

            wb.save(response)
            return response

        else:
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="clinical_pacients.csv"'

            writer = csv.writer(response)
            writer.writerow(columns_names)

            for pacient_data in pacients:
                writer.writerow([
                    pacient_data.id,
                    pacient_data.nsh_id,
                    pacient_data.first_name,
                    pacient_data.last_name,
                    str(pacient_data.date_of_birth),
                    pacient_data.phone,
                    pacient_data.email,
                    pacient_data.default_city,
                    pacient_data.home_addres,

                ])

            return response

def pacientLoad(request):
    if request.method == "POST":
        fileuploadInput = request.FILES['fileuploadInput']

        linkGD = request.POST.get('linkGD')

        gd = GoogleDrive(linkGD)
        gd.request()
        gd.data_from_jjson()


@login_required(login_url='/login/')
def pacientsView(request):
    if request.method=="POST":

        first_name = request.POST.get('first_name')

        last_name = request.POST.get('last_name')

        phone = request.POST.get('phone')

        email = request.POST.get('email')

        nsh_id = request.POST.get('nsh_id')

        default_city = request.POST.get('city')

        home_addres = request.POST.get('addres')
        date_of_birth = request.POST.get('date_of_birth')

        try:
            nsh_id_old = Pacient.objects.all().order_by('-id')[0]
            nsh_id_old = str(nsh_id_old.nsh_id)

            nsh_id_new = gen_aaaannnna(nsh_id_old)
        except:
            nsh_id_new = gen_aaaannnna()



        if email:

            email = vaidator_data.validate_email(str(email))
        else:
            email = ''


        if first_name:
            first_name = vaidator_data.validate_firstname(str(first_name))

        else:
            first_name = ""

        if last_name:
            last_name = vaidator_data.validate_firstname(str(last_name))
        else:
            last_name = ""

        if phone:
            phone = vaidator_data.validate_phone(str(phone))

        else:
            phone = ""

        if default_city:

            default_city = str(default_city)
        else:
            default_city = ''



        if home_addres:

            home_addres = str(home_addres)
        else:
            home_addres = ""

        new_pacient = Pacient.objects.create(
                        email=email,
                        phone=phone,
                        first_name=first_name,
                        last_name=last_name,
                        default_city=default_city,
                        home_addres=home_addres,
                        date_of_birth=date_of_birth
                    )

        new_pacient.nsh_id = nsh_id_new
        new_pacient.save()


        return redirect('/pacients/')


    elif request.method=="GET":

        search_param = {}

        first_name = request.GET.get('first_name')

        last_name = request.GET.get('last_name')

        phone = request.GET.get('phone')

        email = request.GET.get('email')

        nsh_id = request.GET.get('nsh_id')

        default_city = request.GET.get('city')

        home_addres = request.GET.get('addres')
        birth_from = request.GET.get('birth_from')

        birth_to = request.GET.get('birth_to')



        if email:

            email = vaidator_data.validate_email(str(email))

            search_param['email'] = email



        search_param['is_hide'] = False

        if nsh_id:
            nsh_id = str(nsh_id)

            search_param['nsh_id'] = nsh_id

        if first_name:

            search_param['first_name'] = str(first_name)

        if last_name:
            search_param['last_name'] = str(last_name)

        if phone:
            phone = vaidator_data.validate_phone(str(phone))

            search_param['phone'] = phone

        if default_city:

            default_city = str(default_city)

            search_param['default_city'] = default_city

        if home_addres:

            home_addres = str(home_addres)

            search_param['home_addres'] = home_addres

        # if birth_from and birth_to:
        #     birth_from = vaidator_data.validate_date(str(birth_from))
        #
        #     search_param['date_of_birth__range'] = [birth_from, birth_to]
        #
        # elif birth_from:
        #     birth_from = vaidator_data.validate_date(str(birth_from))
        #
        #     search_param['date_of_birth__range'] = [birth_from, '2023-12-30']
        #
        # elif birth_to:
        #     birth_to = vaidator_data.validate_date(str(birth_to))
        #
        #     search_param['date_of_birth__range'] = ['1900-01-01',birth_to]
        #
        # else:
        #     search_param['date_of_birth__range'] = ['1900-01-01','2023-12-30']


        pacients = Pacient.objects.filter(
                        **search_param
                    ).order_by('-id')

        return render(request, 'pacients/pacients.html', {
            'pacients': pacients,
            'res_params':str(request.get_full_path())
                      .replace('/pacients/?','').replace('/pacients/','')
        })



def register_view(request):
    if request.method == "POST":
        email_ = request.POST.get('email')
        username = request.POST.get('username')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')

        phone = request.POST.get('phone')
        role = request.POST.get('role')


        if phone:
            phone = str(phone).replace('+', "")\
                .replace(' ', "").replace('(', "").replace(')', "")\
                .replace('-', "").replace('.', "")


        password = request.POST.get('password')
        repeatpassword = request.POST.get('repeatpassword')

        if password == repeatpassword:

            user_ = User.objects.create_user(
                email=email_,
                username=username,

                password=password,
                first_name=first_name,
                last_name=last_name
            )
            user_.save()

            profile, created = UserProfile.objects.get_or_create(
                user=user_, phone=phone,role=role)
            profile.save()





            user = authenticate(username=username, password=password)
            login(request, user)
        return redirect('/master/')

    elif request.method == "GET":
        return render(request, 'auth/sign_up.html')

run_f_scores_reloading_weig()